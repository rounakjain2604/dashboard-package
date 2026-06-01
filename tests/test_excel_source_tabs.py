from openpyxl import load_workbook

from dashboard_api import _build_config_from_payload
from src.dcf_engine.output.excel_builder import build_excel
from src.dcf_engine.sample_models import get_sample_payload


def test_excel_without_source_data_keeps_existing_tabs(tmp_path):
    payload = get_sample_payload("AAPL")
    cfg = _build_config_from_payload(payload)
    path = build_excel(tmp_path / "plain.xlsx", cfg)
    wb = load_workbook(path, read_only=True, data_only=False)

    assert "Assumptions" in wb.sheetnames
    assert "DCF" in wb.sheetnames
    assert "Source Map" not in wb.sheetnames


def test_excel_with_source_data_adds_premium_tabs(tmp_path):
    payload = get_sample_payload("AAPL")
    cfg = _build_config_from_payload(payload)
    path = build_excel(
        tmp_path / "source.xlsx",
        cfg,
        source_facts=[{"account": "Revenue", "concept": "Revenues", "value": 100, "unit": "USD"}],
        warnings=["Check source coverage."],
        filing_changes=[{"account": "Revenue", "severity": "medium", "valuation_impact": "Review CAGR."}],
        valuation_impacts=[{"severity": "medium", "title": "Revenue moved", "detail": "Review growth."}],
    )
    wb = load_workbook(path, read_only=True, data_only=False)

    assert wb.sheetnames[:6] == [
        "Cover",
        "Assumptions",
        "Source Map",
        "Filing Changes",
        "Valuation Impacts",
        "Warnings",
    ]
    assert "DCF" in wb.sheetnames
