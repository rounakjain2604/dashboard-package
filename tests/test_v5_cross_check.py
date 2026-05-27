"""
V5 Comprehensive Cross-Check Test
==================================
Uses VERY different numbers from both the default config and the V4 test.
Cross-checks every figure between the Python pipeline (dashboard) output
and what the Excel formulas would compute.

Run:  python test_v5_cross_check.py
"""
import sys
if __name__ != '__main__':
    try:
        import pytest
        pytest.skip("Skip script-style test at import time", allow_module_level=True)
    except ImportError:
        pass

import json
import math
import os
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.dcf_engine.config import (
    DCFEngineConfig, CompanyInfo, ForecastConfig, WACCConfig,
    ValuationConfig, MonteCarloConfig, SensitivityConfig,
    CompsConfig, DebtTranche, ScenarioOverrides,
)
from src.dcf_engine.pipeline import run_pipeline

# ══════════════════════════════════════════════════════════════════════
# TEST INPUTS — intentionally odd numbers to catch hardcoded values
# ══════════════════════════════════════════════════════════════════════
BASE_REVENUE   = 7_777_777     # 7.78M
BASE_CASH      = 333_333
BASE_PPE       = 1_111_111
BASE_NWC       = 222_222
BASE_CS        = 444_444
# Auto-computed: RE = Cash + PPE + NWC - Debt - CS
# RE = 333333 + 1111111 + 222222 - 1500000 - 444444 = -277778
# (negative RE is valid — company has accumulated losses)
DEBT_BAL       = 1_500_000
AUTO_RE        = BASE_CASH + BASE_PPE + BASE_NWC - DEBT_BAL - BASE_CS  # -277778

cfg = DCFEngineConfig(
    company=CompanyInfo(name="OddCorp", ticker="ODD", industry="Retail"),
    forecast=ForecastConfig(
        projection_years=5,
        revenue_method="cagr",
        revenue_cagr=0.15,          # 15% growth
        cogs_pct_revenue=0.42,
        sga_pct_revenue=0.18,
        other_opex_pct_revenue=0.03,
        depreciation_rate=0.08,
        amortisation_pct_revenue=0.002,
        capex_method="pct_revenue",
        capex_pct_revenue=0.05,
        capex_fixed=0.0,
        dso=40.0,
        dio=60.0,
        dpo=35.0,
        prepaid_pct_revenue=0.006,
        accrued_pct_revenue=0.015,
        other_current_assets_pct_revenue=0.003,
        other_current_liabilities_pct_revenue=0.007,
        tax_rate=0.28,
        dividend_payout_ratio=0.15,
    ),
    wacc=WACCConfig(
        risk_free_rate=0.045,
        equity_risk_premium=0.058,
        beta=0.85,
        size_premium=0.02,
        country_risk_premium=0.008,
        target_debt_weight=0.40,
        target_equity_weight=0.60,
        interest_coverage_ratio=4.0,  # BB+ spread
        tax_rate=0.28,
        use_live_data=False,
    ),
    valuation=ValuationConfig(
        terminal_growth_rate=0.028,
        exit_ev_ebitda_multiple=8.5,
        gordon_weight=0.45,
        cash=333_333,
        debt=1_500_000,
        fully_diluted_shares=5_000_000,
        gdp_growth_cap=0.030,
        terminal_spread_floor_bps=75.0,
    ),
    monte_carlo=MonteCarloConfig(
        iterations=300,
        revenue_growth_mean=0.15,
        revenue_growth_std=0.05,
        ebitda_margin_mean=0.37,
        ebitda_margin_std=0.07,
        wacc_mean=0.095,
        wacc_std=0.018,
        terminal_growth_mean=0.028,
        terminal_growth_std=0.009,
        exit_multiple_mean=8.5,
        exit_multiple_std=2.0,
        seed=42,
    ),
    sensitivity=SensitivityConfig(),
    comps=CompsConfig(peer_tickers=[]),
    debt_tranches=[
        DebtTranche(
            name="Senior Secured",
            beginning_balance=1_000_000,
            interest_rate=0.065,
            annual_amortisation=100_000,
            maturity_year=5,
        ),
        DebtTranche(
            name="Mezzanine",
            beginning_balance=500_000,
            interest_rate=0.095,
            annual_amortisation=50_000,
            maturity_year=5,
        ),
    ],
    scenarios={
        "Base": ScenarioOverrides(name="Base"),
        "Bull": ScenarioOverrides(name="Bull", revenue_multiplier=1.20,
                                   margin_delta_bps=250),
        "Bear": ScenarioOverrides(name="Bear", revenue_multiplier=0.80,
                                   margin_delta_bps=-300),
    },
)

# ══════════════════════════════════════════════════════════════════════
# RUN PIPELINE
# ══════════════════════════════════════════════════════════════════════
tmp_dir = tempfile.mkdtemp()
excel_path = os.path.join(tmp_dir, "v5_test_model.xlsx")

print("=" * 70)
print("V5 COMPREHENSIVE CROSS-CHECK TEST")
print("=" * 70)
result = run_pipeline(
    cfg=cfg,
    historical=pd.DataFrame(),
    base_year_revenue=BASE_REVENUE,
    base_cash=BASE_CASH,
    base_ppe=BASE_PPE,
    base_nwc=BASE_NWC,
    base_retained_earnings=0,   # Intentionally 0 — pipeline should auto-fix
    base_common_stock=BASE_CS,
    output_excel=excel_path,
)

errors = []
warnings = []

def check(desc, expected, actual, tol=0.01):
    if expected == 0 and actual == 0:
        return True
    if expected == 0:
        diff = abs(actual)
    else:
        diff = abs((actual - expected) / expected)
    if diff > tol:
        msg = f"FAIL: {desc}: expected={expected:,.2f}, got={actual:,.2f}, diff={diff:.4%}"
        errors.append(msg)
        print(f"  X {msg}")
        return False
    else:
        print(f"  OK {desc}: {actual:,.2f}")
        return True

is_df = result.income_statement.table
wc_df = result.working_capital.table
cd_df = result.capex_da.table
ds_df = result.debt_schedule.table
bs_df = result.balance_sheet.table
cf_df = result.cash_flow.table
fcf_df = result.cash_flow.fcf_table

# ══════════════════════════════════════════════════════════════════════
# TEST 1: INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 1: Income Statement ---")
expected_rev = BASE_REVENUE
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    expected_rev = BASE_REVENUE * (1.15 ** yr)
    check(f"Yr{yr} Revenue", expected_rev, row["Revenue"])
    check(f"Yr{yr} COGS = 42% Rev", expected_rev * 0.42, row["COGS"])
    check(f"Yr{yr} SGA = 18% Rev", expected_rev * 0.18, row["SGA"])
    check(f"Yr{yr} OOpex = 3% Rev", expected_rev * 0.03, row["Other OpEx"])
    gp = expected_rev * (1 - 0.42)
    ebitda = gp - expected_rev * 0.18 - expected_rev * 0.03
    check(f"Yr{yr} GP", gp, row["Gross Profit"])
    check(f"Yr{yr} EBITDA", ebitda, row["EBITDA"])
    check(f"Yr{yr} EBIT = EBITDA-D&A", row["EBITDA"] - row["Total D&A"], row["EBIT"])
    check(f"Yr{yr} EBT = EBIT-Int", row["EBIT"] - row["Interest Expense"], row["EBT"])
    check(f"Yr{yr} Tax = max(EBT*28%,0)", max(row["EBT"] * 0.28, 0), row["Tax Expense"])
    check(f"Yr{yr} NI = EBT-Tax", row["EBT"] - row["Tax Expense"], row["Net Income"])

# ══════════════════════════════════════════════════════════════════════
# TEST 2: WORKING CAPITAL
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 2: Working Capital ---")
prev_nwc = BASE_NWC
for i, row in wc_df.iterrows():
    yr = int(row["year_index"])
    rev = is_df.iloc[i]["Revenue"]
    cogs = is_df.iloc[i]["COGS"]
    ar = rev * 40 / 365
    inv = cogs * 60 / 365
    prep = rev * 0.006
    oca = rev * 0.003
    ap = cogs * 35 / 365
    accr = rev * 0.015
    ocl = rev * 0.007
    nwc = ar + inv + prep + oca - ap - accr - ocl
    dnwc = nwc - prev_nwc
    check(f"Yr{yr} WC AR", ar, row["Accounts Receivable"])
    check(f"Yr{yr} WC Inv", inv, row["Inventory"])
    check(f"Yr{yr} WC NWC", nwc, row["NWC"])
    check(f"Yr{yr} WC Delta NWC", dnwc, row["Delta NWC"])
    prev_nwc = nwc

# ══════════════════════════════════════════════════════════════════════
# TEST 3: CAPEX & DEPRECIATION
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 3: Capex & Depreciation ---")
prev_ppe = BASE_PPE
for i, row in cd_df.iterrows():
    yr = int(row["year_index"])
    rev = is_df.iloc[i]["Revenue"]
    capex_exp = rev * 0.05
    dep_ex = prev_ppe * 0.08
    dep_new = capex_exp * 0.08 * 0.5
    dep_total = dep_ex + dep_new
    end_ppe = prev_ppe + capex_exp - dep_total
    check(f"Yr{yr} Beg PP&E", prev_ppe, row["Beginning PP&E (Net)"])
    check(f"Yr{yr} Capex", capex_exp, row["Capex"])
    check(f"Yr{yr} Dep Existing", dep_ex, row["Dep on Existing"])
    check(f"Yr{yr} Dep New", dep_new, row["Dep on New"])
    check(f"Yr{yr} Total Dep", dep_total, row["Depreciation"])
    check(f"Yr{yr} End PP&E", end_ppe, row["Ending PP&E (Net)"])
    prev_ppe = end_ppe

# ══════════════════════════════════════════════════════════════════════
# TEST 4: DEBT SCHEDULE (2 Tranches)
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 4: Debt Schedule ---")
ss_bal = 1_000_000
mz_bal = 500_000
for i, row in ds_df.iterrows():
    yr = int(row["year_index"])
    expected_beg = ss_bal + mz_bal
    ss_int = ss_bal * 0.065
    mz_int = mz_bal * 0.095
    expected_int = ss_int + mz_int
    if yr == 5:
        ss_rep = ss_bal
        mz_rep = mz_bal
    else:
        ss_rep = min(100_000, ss_bal)
        mz_rep = min(50_000, mz_bal)
    expected_rep = ss_rep + mz_rep
    expected_end = max(ss_bal - ss_rep, 0) + max(mz_bal - mz_rep, 0)
    check(f"Yr{yr} Debt Beg", expected_beg, row["Beginning Balance"])
    check(f"Yr{yr} Interest", expected_int, row["Interest Expense"])
    check(f"Yr{yr} Repayment", expected_rep, row["Principal Repayment"])
    check(f"Yr{yr} Debt End", expected_end, row["Ending Balance"])
    ss_bal = max(ss_bal - ss_rep, 0)
    mz_bal = max(mz_bal - mz_rep, 0)

# ══════════════════════════════════════════════════════════════════════
# TEST 5: IS RE-LINKING (Step 5)
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 5: IS Re-linking ---")
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    cd_row = cd_df[cd_df["year_index"] == yr].iloc[0]
    ds_row = ds_df[ds_df["year_index"] == yr].iloc[0]
    check(f"Yr{yr} IS.Dep == CapexDA.Dep", cd_row["Depreciation"], row["Depreciation"])
    check(f"Yr{yr} IS.Interest == Debt.Interest", ds_row["Interest Expense"], row["Interest Expense"])

# ══════════════════════════════════════════════════════════════════════
# TEST 6: CASH FLOW STATEMENT
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 6: Cash Flow ---")
for i, row in cf_df.iterrows():
    yr = int(row["year_index"])
    ni = row["Net Income"]
    da = row["Total D&A"]
    dnwc = row["Change in NWC"]
    cfo = row["CFO"]
    check(f"Yr{yr} CFO = NI+D&A+dNWC", ni + da + dnwc, cfo)
    check(f"Yr{yr} CFI = -Capex", row["CFI"], row["Capex"])
    check(f"Yr{yr} NetChange = CFO+CFI+CFF", cfo + row["CFI"] + row["CFF"], row["Net Change in Cash"])
    check(f"Yr{yr} EndCash = BegCash+Net", row["Beginning Cash"] + row["Net Change in Cash"], row["Ending Cash"])

# ══════════════════════════════════════════════════════════════════════
# TEST 7: BS CASH == CF ENDING CASH (THE V4 BUG)
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 7: BS Cash == CF Ending Cash ---")
for i in range(len(cf_df)):
    yr = int(cf_df.iloc[i]["year_index"])
    cf_cash = cf_df.iloc[i]["Ending Cash"]
    bs_cash = bs_df[bs_df["year_index"] == yr].iloc[0]["Cash & Cash Equivalents"]
    check(f"Yr{yr} CF.EndingCash == BS.Cash", cf_cash, bs_cash)

# ══════════════════════════════════════════════════════════════════════
# TEST 8: BALANCE SHEET BALANCES
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 8: Balance Sheet Balances ---")
for i, row in bs_df.iterrows():
    yr = int(row["year_index"])
    ta = row["Total Assets"]
    tle = row["Total Equity & Liabilities"]
    diff = abs(ta - tle)
    if diff > 1:
        errors.append(f"FAIL: Yr{yr} BS imbalance: A={ta:,.2f} L+E={tle:,.2f}")
        print(f"  X Yr{yr} BS imbalance: A={ta:,.2f} L+E={tle:,.2f}")
    else:
        print(f"  OK Yr{yr} BS balances: A={ta:,.2f} = L+E={tle:,.2f}")

# ══════════════════════════════════════════════════════════════════════
# TEST 9: AUTO-COMPUTED RETAINED EARNINGS
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 9: Auto-Computed RE ---")
# We passed base_retained_earnings=0 but pipeline should have auto-computed
# RE = Cash + PPE + NWC - Debt - CS = 333333 + 1111111 + 222222 - 1500000 - 444444 = -277778
yr1_re = bs_df[bs_df["year_index"] == 1].iloc[0]["Retained Earnings"]
yr1_ni = is_df.iloc[0]["Net Income"]
yr1_div = yr1_ni * 0.15
expected_yr1_re = AUTO_RE + yr1_ni - yr1_div
check(f"Yr1 RE = autoRE({AUTO_RE:,.0f}) + NI - Div", expected_yr1_re, yr1_re)

# ══════════════════════════════════════════════════════════════════════
# TEST 10: WACC
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 10: WACC ---")
w = result.wacc
ke = 0.045 + 0.85 * 0.058 + 0.02 + 0.008  # Rf + Beta*ERP + SP + CRP
check("Ke (CAPM)", ke, w.cost_of_equity)
# ICR=4.0 -> BB+ rating (3.5 <= 4.0 < 4.5) -> spread = 0.0200
from src.dcf_engine.valuation.wacc import synthetic_credit_spread
_rating, _spread = synthetic_credit_spread(4.0)
print(f"  Synthetic rating: {_rating}, spread: {_spread}")
kd_pre = 0.045 + _spread
check("Kd Pre-Tax", kd_pre, w.cost_of_debt_pre_tax)
kd_post = kd_pre * (1 - 0.28)
check("Kd Post-Tax", kd_post, w.cost_of_debt_after_tax)
wacc_exp = 0.60 * ke + 0.40 * kd_post
check("WACC", wacc_exp, w.wacc)

# ══════════════════════════════════════════════════════════════════════
# TEST 11: DCF VALUATION
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 11: DCF Valuation ---")
d = result.dcf
wacc_val = w.wacc

# Discount factors
for i, row in d.valuation_table.iterrows():
    yr = int(row["year_index"])
    expected_df = 1 / ((1 + wacc_val) ** (yr - 0.5))
    check(f"Yr{yr} DF", expected_df, row["Discount Factor"])

# PV sum
pv_sum = sum(d.valuation_table["PV of UFCF"])
check("PV FCF Sum", pv_sum, d.pv_fcf_sum)

# Terminal value
terminal_fcf = float(d.valuation_table.iloc[-1]["Unlevered FCF"])
terminal_ebitda = float(is_df.iloc[-1]["EBITDA"])
check("Terminal FCF", terminal_fcf, d.terminal_fcf)
check("Terminal EBITDA", terminal_ebitda, d.terminal_ebitda)

# effective_g calculation
g = min(0.028, 0.030)   # tg vs gdp_cap
spread_floor = 75 / 10_000  # 75 bps
effective_g = min(g, wacc_val - spread_floor)
gordon_tv = terminal_fcf * (1 + effective_g) / (wacc_val - effective_g)
exit_tv = terminal_ebitda * 8.5
blended_tv = gordon_tv * 0.45 + exit_tv * 0.55
check("Effective TG", effective_g, d.effective_terminal_growth)
check("Gordon TV", gordon_tv, d.gordon_tv)
check("Exit TV", exit_tv, d.exit_tv)
check("Blended TV", blended_tv, d.blended_tv)

# EV
terminal_df = float(d.valuation_table.iloc[-1]["Discount Factor"])
pv_blended = blended_tv * terminal_df
ev_blended = pv_sum + pv_blended
check("PV Blended TV", pv_blended, d.pv_blended_tv)
check("EV Blended", ev_blended, d.ev_blended)

# Equity bridge
equity_blended = ev_blended + 333_333 - 1_500_000
check("Equity Blended", equity_blended, d.equity_blended)

# Price per share
price_blended = equity_blended / 5_000_000
check("Price Blended", price_blended, d.price_blended)

# ══════════════════════════════════════════════════════════════════════
# TEST 12: UFCF CALCULATION
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 12: UFCF ---")
for i, row in fcf_df.iterrows():
    yr = int(row["year_index"])
    ebit = row["EBIT"]
    nopat = ebit * (1 - 0.28)
    da = row["D&A"]
    capex = row["Capex"]
    dnwc = row["Delta NWC"]
    ufcf = nopat + da + capex + dnwc
    check(f"Yr{yr} NOPAT", nopat, row["NOPAT"])
    check(f"Yr{yr} UFCF", ufcf, row["Unlevered FCF"])

# ══════════════════════════════════════════════════════════════════════
# TEST 13: EXCEL FORMULA CROSS-CHECK
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 13: Excel Cross-Check ---")
from openpyxl import load_workbook
wb = load_workbook(excel_path)
from src.dcf_engine.output.excel_formats import A
from src.dcf_engine.output.sheets_core import IS as IS_R, WC as WC_R, CD as CD_R, DS as DS_R, BS as BS_R, CF as CF_R

# Assumptions sheet values
ws_a = wb["Assumptions"]
checks_13 = {
    "rev_cagr": 0.15,
    "cogs_pct": 0.42,
    "sga_pct": 0.18,
    "oopex_pct": 0.03,
    "dep_rate": 0.08,
    "capex_pct": 0.05,
    "tax": 0.28,
    "dso": 40.0,
    "dio": 60.0,
    "dpo": 35.0,
    "div": 0.15,
    "rf": 0.045,
    "erp": 0.058,
    "beta": 0.85,
    "sp": 0.02,
    "crp": 0.008,
    "dw": 0.40,
    "ew": 0.60,
    "icr": 4.0,
    "tg": 0.028,
    "exit_m": 8.5,
    "gw": 0.45,
    "shares": 5_000_000,
    "cash": 333_333,
    "debt": 1_500_000,
    "brev": 7_777_777,
    "bcash": 333_333,
    "bppe": 1_111_111,
    "bnwc": 222_222,
    "bcs": 444_444,
}
for key, expected in checks_13.items():
    row = A[key]
    val = ws_a.cell(row, 2).value
    if val is None:
        errors.append(f"FAIL: Assumptions[{key}] (row {row}) is None")
        print(f"  X Assumptions[{key}] (row {row}) is None")
    elif isinstance(val, str) and val.startswith("="):
        print(f"  ~ Assumptions[{key}] (row {row}) = formula: {val}")
    elif isinstance(val, (int, float)):
        if abs(val - expected) > 0.001:
            errors.append(f"FAIL: Assumptions[{key}] = {val}, expected {expected}")
            print(f"  X Assumptions[{key}] = {val}, expected {expected}")
        else:
            print(f"  OK Assumptions[{key}] = {val}")

# Check bre formula
bre_val = ws_a.cell(A["bre"], 2).value
if isinstance(bre_val, str) and bre_val.startswith("="):
    print(f"  OK Assumptions[bre] = formula: {bre_val}")
else:
    errors.append(f"FAIL: Assumptions[bre] should be formula, got: {bre_val}")
    print(f"  X Assumptions[bre] should be formula, got: {bre_val}")

# Check IS formulas
ws_is = wb["Income Statement"]
yr1_rev = ws_is.cell(IS_R["Rev"], 2).value
if isinstance(yr1_rev, str) and "$B$43" in yr1_rev and "$B$4" in yr1_rev:
    print(f"  OK IS Yr1 Rev formula refs brev and rev_cagr")
else:
    errors.append(f"FAIL: IS Rev formula: {yr1_rev}")
    print(f"  X IS Rev formula: {yr1_rev}")

# Check BS Cash = CF EndC
ws_bs = wb["Balance Sheet"]
ws_cf = wb["Cash Flow"]
yr1_bs_cash = ws_bs.cell(BS_R["Cash"], 2).value
if isinstance(yr1_bs_cash, str) and "Cash Flow" in yr1_bs_cash:
    print(f"  OK BS Cash Yr1 refs Cash Flow sheet")
else:
    errors.append(f"FAIL: BS Cash formula: {yr1_bs_cash}")
    print(f"  X BS Cash formula: {yr1_bs_cash}")

# Check CF EndC formula
yr1_cf_end = ws_cf.cell(CF_R["EndC"], 2).value
if isinstance(yr1_cf_end, str) and "BegC" in str(CF_R["BegC"]):
    print(f"  OK CF EndC Yr1 = BegC + Net (formula: {yr1_cf_end})")
else:
    print(f"  ~ CF EndC Yr1 formula: {yr1_cf_end}")

# Check BS RE refs bre
yr1_bs_re = ws_bs.cell(BS_R["RE"], 2).value
if isinstance(yr1_bs_re, str) and f"$B${A['bre']}" in yr1_bs_re:
    print(f"  OK BS RE Yr1 refs Assumptions bre")
else:
    errors.append(f"FAIL: BS RE formula: {yr1_bs_re}")
    print(f"  X BS RE formula: {yr1_bs_re}")

# Check Checks tab
ws_chk = wb["Checks"]
yr1_chk_bs = ws_chk.cell(3, 2).value
if isinstance(yr1_chk_bs, str) and "Balance Sheet" in yr1_chk_bs:
    print(f"  OK Checks BS ref Balance Sheet")
yr1_chk_cf = ws_chk.cell(6, 2).value
if isinstance(yr1_chk_cf, str) and "Cash Flow" in yr1_chk_cf:
    print(f"  OK Checks CF ref Cash Flow and Balance Sheet")

# ══════════════════════════════════════════════════════════════════════
# TEST 14: SCENARIO COMPARISON
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 14: Scenario Comparison ---")
if result.scenario_comparison:
    comp = result.scenario_comparison.comparison
    if "EV (Blended)" in comp.index:
        base_ev = comp.loc["EV (Blended)", "Base"]
        bull_ev = comp.loc["EV (Blended)", "Bull"]
        bear_ev = comp.loc["EV (Blended)", "Bear"]
        if bull_ev <= base_ev:
            errors.append(f"FAIL: Bull EV ({bull_ev:,.0f}) should be > Base EV ({base_ev:,.0f})")
            print(f"  X Bull EV should be > Base EV")
        else:
            print(f"  OK Bull EV ({bull_ev:,.0f}) > Base EV ({base_ev:,.0f})")
        if bear_ev >= base_ev:
            errors.append(f"FAIL: Bear EV ({bear_ev:,.0f}) should be < Base EV ({base_ev:,.0f})")
            print(f"  X Bear EV should be < Base EV")
        else:
            print(f"  OK Bear EV ({bear_ev:,.0f}) < Base EV ({base_ev:,.0f})")

# ══════════════════════════════════════════════════════════════════════
# TEST 15: TORNADO
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 15: Tornado ---")
if result.tornado:
    check("Tornado base_equity == DCF equity", d.equity_blended, result.tornado.base_equity)

# ══════════════════════════════════════════════════════════════════════
# TEST 16: DEBT SCHEDULE MULTI-TRANCHE CONSOLIDATION
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 16: Multi-Tranche Consolidation ---")
# Total beginning balance should be sum of both tranches
check("Yr1 Total Debt Beg", 1_500_000, ds_df.iloc[0]["Beginning Balance"])
# Total interest Yr1 = 1M * 6.5% + 500k * 9.5% = 65000 + 47500 = 112500
check("Yr1 Total Interest", 112_500, ds_df.iloc[0]["Interest Expense"])

# ══════════════════════════════════════════════════════════════════════
# TEST 17: EXCEL FORMULA VALUES MATCH PYTHON
# ══════════════════════════════════════════════════════════════════════
print("\n--- TEST 17: Excel Formula Values vs Python ---")
# Manually evaluate what Excel formulas would compute for Year 1
ex_rev_y1 = BASE_REVENUE * (1 + 0.15)
py_rev_y1 = is_df.iloc[0]["Revenue"]
check("Yr1 Rev: Excel vs Python", ex_rev_y1, py_rev_y1)

ex_cogs_y1 = ex_rev_y1 * 0.42
check("Yr1 COGS: Excel vs Python", ex_cogs_y1, is_df.iloc[0]["COGS"])

ex_dep_y1 = BASE_PPE * 0.08 + (ex_rev_y1 * 0.05) * 0.08 * 0.5
check("Yr1 Dep: Excel vs Python", ex_dep_y1, cd_df.iloc[0]["Depreciation"])

# WACC
ex_ke = 0.045 + 0.85 * 0.058 + 0.02 + 0.008
check("Ke: Excel vs Python", ex_ke, w.cost_of_equity)

# ══════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("FINAL SUMMARY")
print("=" * 70)

if errors:
    print(f"\n  {len(errors)} ERRORS FOUND:")
    for e in errors:
        print(f"  - {e}")
else:
    print(f"\n  NO ERRORS FOUND - ALL CHECKS PASSED")

if warnings:
    print(f"\n  {len(warnings)} WARNINGS:")
    for w_msg in warnings:
        print(f"  - {w_msg}")

print(f"\nExcel file: {excel_path}")
print(f"Pipeline errors: {result.errors}")
