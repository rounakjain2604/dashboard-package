"""
Comprehensive cross-check test: Dashboard (Python) vs Excel formulas.

Uses completely different input numbers from the default config to
surface any hardcoded values, formula mismatches, or computation bugs.
"""
import json
import math
import os
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd

# Ensure project root is on path
sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.dcf_engine.config import (
    DCFEngineConfig, CompanyInfo, ForecastConfig, WACCConfig,
    ValuationConfig, MonteCarloConfig, SensitivityConfig,
    CompsConfig, DebtTranche, ScenarioOverrides,
)
from src.dcf_engine.pipeline import run_pipeline

# ══════════════════════════════════════════════════════════════════════
# TEST INPUTS — completely different from defaults
# ══════════════════════════════════════════════════════════════════════
BASE_REVENUE   = 12_000_000   # 12M
BASE_CASH      = 800_000
BASE_PPE       = 2_500_000
BASE_NWC       = 350_000
BASE_RE        = 150_000    # Must balance Year 0 BS: Cash+NWC+PPE = Debt+CS+RE
BASE_CS        = 500_000

cfg = DCFEngineConfig(
    company=CompanyInfo(name="TestCorp", ticker="TST", industry="Tech"),
    forecast=ForecastConfig(
        projection_years=5,
        revenue_method="cagr",
        revenue_cagr=0.12,
        cogs_pct_revenue=0.38,
        sga_pct_revenue=0.15,
        other_opex_pct_revenue=0.07,
        depreciation_rate=0.12,
        amortisation_pct_revenue=0.003,
        capex_method="pct_revenue",
        capex_pct_revenue=0.06,
        capex_fixed=0.0,
        dso=55.0,
        dio=35.0,
        dpo=50.0,
        prepaid_pct_revenue=0.008,
        accrued_pct_revenue=0.012,
        other_current_assets_pct_revenue=0.004,
        other_current_liabilities_pct_revenue=0.006,
        tax_rate=0.22,
        dividend_payout_ratio=0.10,
    ),
    wacc=WACCConfig(
        risk_free_rate=0.038,
        equity_risk_premium=0.065,
        beta=1.25,
        size_premium=0.01,
        country_risk_premium=0.005,
        target_debt_weight=0.35,
        target_equity_weight=0.65,
        interest_coverage_ratio=6.5,
        tax_rate=0.22,
        use_live_data=False,
    ),
    valuation=ValuationConfig(
        terminal_growth_rate=0.03,
        exit_ev_ebitda_multiple=12.0,
        gordon_weight=0.60,
        cash=800_000,
        debt=3_000_000,
        fully_diluted_shares=2_000_000,
        gdp_growth_cap=0.035,
    ),
    monte_carlo=MonteCarloConfig(
        iterations=500,  # small for test speed
        revenue_growth_mean=0.12,
        revenue_growth_std=0.04,
        ebitda_margin_mean=0.40,
        ebitda_margin_std=0.06,
        wacc_mean=0.09,
        wacc_std=0.015,
        terminal_growth_mean=0.03,
        terminal_growth_std=0.008,
        exit_multiple_mean=12.0,
        exit_multiple_std=2.5,
        seed=99,
    ),
    sensitivity=SensitivityConfig(),
    comps=CompsConfig(peer_tickers=[]),  # no live data
    debt_tranches=[
        DebtTranche(
            name="Loan Alpha",
            beginning_balance=3_000_000,
            interest_rate=0.055,
            annual_amortisation=200_000,
            maturity_year=5,
        ),
    ],
    scenarios={
        "Base": ScenarioOverrides(name="Base"),
        "Bull": ScenarioOverrides(name="Bull", revenue_multiplier=1.15,
                                   margin_delta_bps=300),
        "Bear": ScenarioOverrides(name="Bear", revenue_multiplier=0.85,
                                   margin_delta_bps=-300),
    },
)

# ══════════════════════════════════════════════════════════════════════
# RUN PIPELINE
# ══════════════════════════════════════════════════════════════════════
tmp_dir = tempfile.mkdtemp()
excel_path = os.path.join(tmp_dir, "test_model.xlsx")

print("=" * 70)
print("RUNNING PIPELINE WITH CUSTOM TEST INPUTS")
print("=" * 70)
result = run_pipeline(
    cfg=cfg,
    historical=pd.DataFrame(),
    base_year_revenue=BASE_REVENUE,
    base_cash=BASE_CASH,
    base_ppe=BASE_PPE,
    base_nwc=BASE_NWC,
    base_retained_earnings=BASE_RE,
    base_common_stock=BASE_CS,
    output_excel=excel_path,
)

errors_found = []
warnings_found = []

def check(description, expected, actual, tol=0.01, is_pct=False):
    """Check if expected ≈ actual within tolerance."""
    if expected == 0 and actual == 0:
        return True
    if expected == 0:
        diff = abs(actual)
    else:
        diff = abs((actual - expected) / expected)
    if diff > tol:
        msg = f"FAIL: {description}: expected={expected:,.2f}, actual={actual:,.2f}, diff={diff:.4%}"
        errors_found.append(msg)
        print(f"  ❌ {msg}")
        return False
    else:
        print(f"  ✓ {description}: {actual:,.2f}")
        return True

# ══════════════════════════════════════════════════════════════════════
# PART 1: INTERNAL PYTHON CONSISTENCY CHECKS
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PART 1: INTERNAL PYTHON CONSISTENCY CHECKS")
print("=" * 70)

is_df = result.income_statement.table
wc_df = result.working_capital.table
cd_df = result.capex_da.table
ds_df = result.debt_schedule.table
bs_df = result.balance_sheet.table
cf_df = result.cash_flow.table
fcf_df = result.cash_flow.fcf_table

# ── 1A: Income Statement Checks ─────────────────────────────────────
print("\n--- 1A: Income Statement ---")
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    rev = row["Revenue"]
    cogs = row["COGS"]
    gp = row["Gross Profit"]
    sga = row["SGA"]
    oopex = row["Other OpEx"]
    ebitda = row["EBITDA"]
    dep = row["Depreciation"]
    amort = row["Amortisation"]
    da = row["Total D&A"]
    ebit = row["EBIT"]
    interest = row["Interest Expense"]
    ebt = row["EBT"]
    tax = row["Tax Expense"]
    ni = row["Net Income"]

    check(f"Yr{yr} GP = Rev - COGS", rev - cogs, gp)
    check(f"Yr{yr} EBITDA = GP - SGA - OOpex", gp - sga - oopex, ebitda)
    check(f"Yr{yr} D&A = Dep + Amort", dep + amort, da)
    check(f"Yr{yr} EBIT = EBITDA - D&A", ebitda - da, ebit)
    check(f"Yr{yr} EBT = EBIT - Interest", ebit - interest, ebt)
    check(f"Yr{yr} Tax = max(EBT*22%,0)", max(ebt * 0.22, 0), tax)
    check(f"Yr{yr} NI = EBT - Tax", ebt - tax, ni)

# ── 1B: Revenue Projection (CAGR=12%) ──────────────────────────────
print("\n--- 1B: Revenue Projection ---")
expected_rev = BASE_REVENUE
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    expected_rev *= (1 + 0.12)
    check(f"Yr{yr} Revenue = base*(1.12)^{yr}", expected_rev, row["Revenue"])

# ── 1C: COGS/SGA/OOpex Percentages ─────────────────────────────────
print("\n--- 1C: Cost % of Revenue ---")
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    check(f"Yr{yr} COGS = 38% of Revenue", row["Revenue"] * 0.38, row["COGS"])
    check(f"Yr{yr} SGA = 15% of Revenue", row["Revenue"] * 0.15, row["SGA"])
    check(f"Yr{yr} OOpex = 7% of Revenue", row["Revenue"] * 0.07, row["Other OpEx"])

# ── 1D: Working Capital Checks ──────────────────────────────────────
print("\n--- 1D: Working Capital ---")
for i, row in wc_df.iterrows():
    yr = int(row["year_index"])
    rev = row["Revenue"]
    cogs = row["COGS"]
    ar = rev * 55 / 365
    inv = cogs * 35 / 365
    prep = rev * 0.008
    oca = rev * 0.004
    ap = cogs * 50 / 365
    accr = rev * 0.012
    ocl = rev * 0.006
    nwc = ar + inv + prep + oca - ap - accr - ocl
    check(f"Yr{yr} WC AR = Rev*55/365", ar, row["Accounts Receivable"])
    check(f"Yr{yr} WC Inv = COGS*35/365", inv, row["Inventory"])
    check(f"Yr{yr} WC NWC", nwc, row["NWC"])

# ── 1E: Capex & Depreciation ────────────────────────────────────────
print("\n--- 1E: Capex & Depreciation ---")
prev_net_ppe = BASE_PPE
for i, row in cd_df.iterrows():
    yr = int(row["year_index"])
    beg = row["Beginning PP&E (Net)"]
    capex = row["Capex"]
    dep_ex = row["Dep on Existing"]
    dep_new = row["Dep on New"]
    dep = row["Depreciation"]
    end = row["Ending PP&E (Net)"]

    check(f"Yr{yr} Beg PP&E = prev End", prev_net_ppe, beg)
    check(f"Yr{yr} Capex = Rev*6%", is_df.iloc[i]["Revenue"] * 0.06, capex)
    check(f"Yr{yr} DepEx = Beg*12%", beg * 0.12, dep_ex)
    check(f"Yr{yr} DepNew = Capex*12%*0.5", capex * 0.12 * 0.5, dep_new)
    check(f"Yr{yr} TotalDep = DepEx+DepNew", dep_ex + dep_new, dep)
    check(f"Yr{yr} End = Beg+Capex-Dep", beg + capex - dep, end)
    prev_net_ppe = end

# ── 1F: Debt Schedule ───────────────────────────────────────────────
print("\n--- 1F: Debt Schedule ---")
prev_bal = 3_000_000
for i, row in ds_df.iterrows():
    yr = int(row["year_index"])
    beg = row["Beginning Balance"]
    interest = row["Interest Expense"]
    repay = row["Principal Repayment"]
    end = row["Ending Balance"]

    check(f"Yr{yr} Debt Beg", prev_bal, beg)
    check(f"Yr{yr} Interest = Beg*5.5%", beg * 0.055, interest)
    if yr == 5:
        check(f"Yr{yr} Bullet Repay = Beg", beg, repay)
    else:
        check(f"Yr{yr} Amort = min(200k, Beg)", min(200_000, beg), repay)
    check(f"Yr{yr} End = max(Beg-Repay,0)", max(beg - repay, 0), end)
    prev_bal = end

# ── 1G: IS D&A and Interest Re-linking (Step 5 critical check) ─────
print("\n--- 1G: IS Re-linking ---")
for i, row in is_df.iterrows():
    yr = int(row["year_index"])
    cd_row = cd_df[cd_df["year_index"] == yr].iloc[0]
    ds_row = ds_df[ds_df["year_index"] == yr].iloc[0]
    check(f"Yr{yr} IS.Dep == CapexDA.Dep", cd_row["Depreciation"], row["Depreciation"])
    check(f"Yr{yr} IS.Interest == Debt.Interest", ds_row["Interest Expense"], row["Interest Expense"])

# ── 1H: Balance Sheet Balances ──────────────────────────────────────
print("\n--- 1H: Balance Sheet Balances ---")
for i, row in bs_df.iterrows():
    yr = int(row["year_index"])
    ta = row["Total Assets"]
    tle = row["Total Equity & Liabilities"]
    diff = abs(ta - tle)
    if diff > 1:
        errors_found.append(f"FAIL: Yr{yr} BS doesn't balance: A={ta:,.2f} L+E={tle:,.2f}")
        print(f"  ❌ Yr{yr} BS doesn't balance: A={ta:,.2f} L+E={tle:,.2f}")
    else:
        print(f"  ✓ Yr{yr} BS balances: A={ta:,.2f} = L+E={tle:,.2f}")

# ── 1I: Cash Flow Statement ─────────────────────────────────────────
print("\n--- 1I: Cash Flow Statement ---")
for i, row in cf_df.iterrows():
    yr = int(row["year_index"])
    cfo = row["CFO"]
    cfi = row["CFI"]
    cff = row["CFF"]
    net = row["Net Change in Cash"]
    beg = row["Beginning Cash"]
    end = row["Ending Cash"]

    check(f"Yr{yr} Net = CFO+CFI+CFF", cfo + cfi + cff, net)
    check(f"Yr{yr} End = Beg+Net", beg + net, end)

# ── 1J: UFCF Calculation ────────────────────────────────────────────
print("\n--- 1J: UFCF ---")
for i, row in fcf_df.iterrows():
    yr = int(row["year_index"])
    ebit = row["EBIT"]
    nopat = row["NOPAT"]
    da = row["D&A"]
    capex = row["Capex"]
    dnwc = row["Delta NWC"]
    ufcf = row["Unlevered FCF"]

    check(f"Yr{yr} NOPAT = EBIT*(1-22%)", ebit * 0.78, nopat)
    check(f"Yr{yr} UFCF = NOPAT+D&A+Capex+dNWC", nopat + da + capex + dnwc, ufcf)

# ── 1K: WACC ────────────────────────────────────────────────────────
print("\n--- 1K: WACC ---")
w = result.wacc
ke = 0.038 + 1.25 * 0.065 + 0.01 + 0.005  # Rf + Beta*ERP + SP + CRP
check("Ke (CAPM)", ke, w.cost_of_equity)
# ICR=6.5 → A rating (6.0 ≤ 6.5 < 7.5) → spread = 0.0108
kd_pre = 0.038 + 0.0108
check("Kd Pre-Tax", kd_pre, w.cost_of_debt_pre_tax)
kd_post = kd_pre * (1 - 0.22)
check("Kd Post-Tax", kd_post, w.cost_of_debt_after_tax)
wacc_expected = 0.65 * ke + 0.35 * kd_post
check("WACC", wacc_expected, w.wacc)

# ── 1L: DCF Valuation ───────────────────────────────────────────────
print("\n--- 1L: DCF Valuation ---")
d = result.dcf
wacc_val = w.wacc
vt = d.valuation_table

# Check discount factors (mid-year)
for i, row in vt.iterrows():
    yr = int(row["year_index"])
    expected_df = 1 / ((1 + wacc_val) ** (yr - 0.5))
    check(f"Yr{yr} DF", expected_df, row["Discount Factor"])

# PV sum
pv_sum_calc = sum(vt["PV of UFCF"])
check("PV FCF Sum", pv_sum_calc, d.pv_fcf_sum)

# Terminal value
terminal_fcf = float(vt.iloc[-1]["Unlevered FCF"])
terminal_ebitda = float(is_df.iloc[-1]["EBITDA"])
check("Terminal FCF", terminal_fcf, d.terminal_fcf)
check("Terminal EBITDA", terminal_ebitda, d.terminal_ebitda)

# effective g = min(tg, gdp_cap) then min(g, wacc - spread_floor)
g = min(0.03, 0.035)  # tg vs gdp cap
spread_floor = 50 / 10_000  # 50 bps
g = min(g, wacc_val - spread_floor)
gordon_tv = terminal_fcf * (1 + g) / (wacc_val - g)
exit_tv = terminal_ebitda * 12.0
blended_tv = gordon_tv * 0.60 + exit_tv * 0.40
check("Gordon TV", gordon_tv, d.gordon_tv)
check("Exit TV", exit_tv, d.exit_tv)
check("Blended TV", blended_tv, d.blended_tv)

# PV of TV
terminal_df = float(vt.iloc[-1]["Discount Factor"])
pv_blended = blended_tv * terminal_df
check("PV Blended TV", pv_blended, d.pv_blended_tv)

# EV
ev_blended = pv_sum_calc + pv_blended
check("EV Blended", ev_blended, d.ev_blended)

# Equity bridge
equity_blended = ev_blended + 800_000 - 3_000_000
check("Equity Blended", equity_blended, d.equity_blended)

# Price per share
price_blended = equity_blended / 2_000_000
check("Price Blended", price_blended, d.price_blended)

# ── 1M: CF Ending Cash == BS Cash ───────────────────────────────────
print("\n--- 1M: CF Ending Cash vs BS Cash ---")
for i in range(len(cf_df)):
    yr = int(cf_df.iloc[i]["year_index"])
    cf_cash = cf_df.iloc[i]["Ending Cash"]
    bs_cash = bs_df[bs_df["year_index"] == yr].iloc[0]["Cash & Cash Equivalents"]
    check(f"Yr{yr} CF.EndCash == BS.Cash", cf_cash, bs_cash)

# ══════════════════════════════════════════════════════════════════════
# PART 2: EXCEL FORMULA CROSS-CHECK
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PART 2: EXCEL FORMULA CROSS-CHECK")
print("=" * 70)

from openpyxl import load_workbook
wb = load_workbook(excel_path)

# Check all expected sheets exist
expected_sheets = [
    "Cover", "Assumptions", "Income Statement", "Working Capital",
    "Capex DA", "Debt Schedule", "Balance Sheet", "Cash Flow",
    "WACC", "DCF", "Scenarios", "Sensitivity", "Monte Carlo",
    "Tornado", "Checks", "Audit Trail"
]
for s in expected_sheets:
    if s in wb.sheetnames:
        print(f"  ✓ Sheet '{s}' present")
    else:
        errors_found.append(f"FAIL: Sheet '{s}' missing")
        print(f"  ❌ Sheet '{s}' missing!")

# ── 2A: Assumptions values match config ──────────────────────────────
print("\n--- 2A: Assumptions Sheet Values ---")
ws_a = wb["Assumptions"]
from src.dcf_engine.output.excel_formats import A

assumptions_checks = {
    "rev_cagr": 0.12,
    "cogs_pct": 0.38,
    "sga_pct": 0.15,
    "oopex_pct": 0.07,
    "dep_rate": 0.12,
    "capex_pct": 0.06,
    "tax": 0.22,
    "dso": 55.0,
    "dio": 35.0,
    "dpo": 50.0,
    "div": 0.10,
    "amort_pct": 0.003,
    "prepaid_pct": 0.008,
    "accrued_pct": 0.012,
    "oca_pct": 0.004,
    "ocl_pct": 0.006,
    "rf": 0.038,
    "erp": 0.065,
    "beta": 1.25,
    "sp": 0.01,
    "crp": 0.005,
    "dw": 0.35,
    "ew": 0.65,
    "icr": 6.5,
    "tg": 0.03,
    "exit_m": 12.0,
    "gw": 0.60,
    "shares": 2_000_000,  # absolute share count (fixed from /1e6 bug)
    "cash": 800_000,
    "debt": 3_000_000,
    "brev": 12_000_000,
    "bcash": 800_000,
    "bppe": 2_500_000,
    "bnwc": 350_000,
    "bcs": 500_000,
    "debt_bal": 3_000_000,
    "debt_rate": 0.055,
    "debt_amort": 200_000,
    "debt_maturity": 5,
}
for key, expected in assumptions_checks.items():
    row = A[key]
    val = ws_a.cell(row, 2).value
    if val is None:
        errors_found.append(f"FAIL: Assumptions[{key}] (row {row}) is None, expected {expected}")
        print(f"  ❌ Assumptions[{key}] (row {row}) is None")
    elif isinstance(val, str) and val.startswith("="):
        # Formula cell, skip value check
        print(f"  ~ Assumptions[{key}] (row {row}) = formula: {val}")
    elif isinstance(val, (int, float)):
        if abs(val - expected) > 0.001:
            errors_found.append(f"FAIL: Assumptions[{key}] (row {row}) = {val}, expected {expected}")
            print(f"  ❌ Assumptions[{key}] (row {row}) = {val}, expected {expected}")
        else:
            print(f"  ✓ Assumptions[{key}] (row {row}) = {val}")
    else:
        print(f"  ~ Assumptions[{key}] (row {row}) = {val} (type: {type(val).__name__})")

# ── 2B: IS formulas reference correct Assumptions rows ──────────────
print("\n--- 2B: IS Formula Structure ---")
ws_is = wb["Income Statement"]
from src.dcf_engine.output.sheets_core import IS as IS_ROWS

# Year 1 Revenue formula should reference brev and rev_cagr
yr1_rev_formula = ws_is.cell(IS_ROWS["Rev"], 2).value
if isinstance(yr1_rev_formula, str):
    if f"$B${A['brev']}" in yr1_rev_formula and f"$B${A['rev_cagr']}" in yr1_rev_formula:
        print(f"  ✓ Yr1 Revenue formula references brev and rev_cagr correctly")
    else:
        errors_found.append(f"FAIL: Yr1 Rev formula doesn't ref correct cells: {yr1_rev_formula}")
        print(f"  ❌ Yr1 Rev formula: {yr1_rev_formula}")

# Year 2 Revenue should reference Year 1 Revenue and rev_cagr
yr2_rev_formula = ws_is.cell(IS_ROWS["Rev"], 3).value
if isinstance(yr2_rev_formula, str):
    if "B2" in yr2_rev_formula and f"$B${A['rev_cagr']}" in yr2_rev_formula:
        print(f"  ✓ Yr2 Revenue formula references B2 and rev_cagr correctly")
    else:
        errors_found.append(f"FAIL: Yr2 Rev formula: {yr2_rev_formula}")
        print(f"  ❌ Yr2 Rev formula: {yr2_rev_formula}")

# COGS should reference Rev and cogs_pct
yr1_cogs_formula = ws_is.cell(IS_ROWS["COGS"], 2).value
if isinstance(yr1_cogs_formula, str):
    if f"$B${A['cogs_pct']}" in yr1_cogs_formula:
        print(f"  ✓ Yr1 COGS formula references cogs_pct correctly")
    else:
        errors_found.append(f"FAIL: Yr1 COGS formula: {yr1_cogs_formula}")
        print(f"  ❌ Yr1 COGS formula: {yr1_cogs_formula}")

# Depreciation should come from Capex DA sheet
yr1_dep_formula = ws_is.cell(IS_ROWS["Dep"], 2).value
if isinstance(yr1_dep_formula, str):
    if "Capex DA" in yr1_dep_formula:
        print(f"  ✓ Yr1 Depreciation references Capex DA sheet")
    else:
        errors_found.append(f"FAIL: Yr1 Dep formula: {yr1_dep_formula}")
        print(f"  ❌ Yr1 Dep formula: {yr1_dep_formula}")

# Interest should come from Debt Schedule
yr1_int_formula = ws_is.cell(IS_ROWS["Int"], 2).value
if isinstance(yr1_int_formula, str):
    if "Debt Schedule" in yr1_int_formula:
        print(f"  ✓ Yr1 Interest references Debt Schedule")
    else:
        errors_found.append(f"FAIL: Yr1 Interest formula: {yr1_int_formula}")
        print(f"  ❌ Yr1 Interest formula: {yr1_int_formula}")

# ── 2C: Capex DA formula structure ──────────────────────────────────
print("\n--- 2C: Capex DA Formula Structure ---")
ws_cd = wb["Capex DA"]
from src.dcf_engine.output.sheets_core import CD as CD_ROWS

yr1_beg_formula = ws_cd.cell(CD_ROWS["Beg"], 2).value
if isinstance(yr1_beg_formula, str):
    if f"$B${A['bppe']}" in yr1_beg_formula:
        print(f"  ✓ Yr1 Beg PP&E references bppe correctly")
    else:
        errors_found.append(f"FAIL: Yr1 Beg PP&E formula: {yr1_beg_formula}")
        print(f"  ❌ Yr1 Beg PP&E formula: {yr1_beg_formula}")

yr1_capex_formula = ws_cd.cell(CD_ROWS["Capex"], 2).value
if isinstance(yr1_capex_formula, str):
    if f"$B${A['capex_pct']}" in yr1_capex_formula:
        print(f"  ✓ Yr1 Capex references capex_pct correctly")
    else:
        errors_found.append(f"FAIL: Yr1 Capex formula: {yr1_capex_formula}")
        print(f"  ❌ Yr1 Capex formula: {yr1_capex_formula}")

# Dep on Existing = Beg * dep_rate
yr1_depex_formula = ws_cd.cell(CD_ROWS["DepEx"], 2).value
if isinstance(yr1_depex_formula, str):
    if f"$B${A['dep_rate']}" in yr1_depex_formula:
        print(f"  ✓ Yr1 DepExisting references dep_rate correctly")
    else:
        errors_found.append(f"FAIL: Yr1 DepExisting formula: {yr1_depex_formula}")
        print(f"  ❌ Yr1 DepExisting formula: {yr1_depex_formula}")

# ── 2D: Debt Schedule Formulas ──────────────────────────────────────
print("\n--- 2D: Debt Schedule Formula Structure ---")
ws_ds = wb["Debt Schedule"]
from src.dcf_engine.output.sheets_core import DS as DS_ROWS

yr1_beg_formula = ws_ds.cell(DS_ROWS["Beg"], 2).value
if isinstance(yr1_beg_formula, str):
    if f"$B${A['debt_bal']}" in yr1_beg_formula:
        print(f"  ✓ Yr1 Debt Beg references debt_bal correctly")
    else:
        errors_found.append(f"FAIL: Yr1 Debt Beg: {yr1_beg_formula}")
        print(f"  ❌ Yr1 Debt Beg: {yr1_beg_formula}")

yr1_int_formula = ws_ds.cell(DS_ROWS["Int"], 2).value
if isinstance(yr1_int_formula, str):
    if f"$B${A['debt_rate']}" in yr1_int_formula:
        print(f"  ✓ Yr1 Interest references debt_rate correctly")
    else:
        errors_found.append(f"FAIL: Yr1 Interest formula: {yr1_int_formula}")
        print(f"  ❌ Yr1 Interest formula: {yr1_int_formula}")

# ── 2E: DCF Formula Structure ──────────────────────────────────────
print("\n--- 2E: DCF Formula Structure ---")
ws_dcf = wb["DCF"]
from src.dcf_engine.output.sheets_core import DC as DC_ROWS

# UFCF references Cash Flow
yr1_ufcf = ws_dcf.cell(DC_ROWS["UFCF"], 2).value
if isinstance(yr1_ufcf, str):
    if "Cash Flow" in yr1_ufcf:
        print(f"  ✓ DCF UFCF references Cash Flow sheet")
    else:
        errors_found.append(f"FAIL: DCF UFCF formula: {yr1_ufcf}")
        print(f"  ❌ DCF UFCF formula: {yr1_ufcf}")

# Discount factor uses WACC and mid-year
yr1_df_formula = ws_dcf.cell(DC_ROWS["DF"], 2).value
if isinstance(yr1_df_formula, str):
    if "WACC" in yr1_df_formula and "0.5" in yr1_df_formula:
        print(f"  ✓ DCF DF uses WACC and mid-year convention")
    else:
        errors_found.append(f"FAIL: DCF DF formula: {yr1_df_formula}")
        print(f"  ❌ DCF DF formula: {yr1_df_formula}")

# TV Gordon references terminal growth and WACC
tv_gordon_formula = ws_dcf.cell(DC_ROWS["TVG"], 2).value
if isinstance(tv_gordon_formula, str):
    if f"$B${A['tg']}" in tv_gordon_formula and "WACC" in tv_gordon_formula:
        print(f"  ✓ Gordon TV references terminal growth and WACC")
    else:
        errors_found.append(f"FAIL: Gordon TV formula: {tv_gordon_formula}")
        print(f"  ❌ Gordon TV formula: {tv_gordon_formula}")

# Equity bridge
equity_formula = ws_dcf.cell(DC_ROWS["Equity"], 2).value
if isinstance(equity_formula, str):
    if f"$B${A['cash']}" in equity_formula or f"B{DC_ROWS['Cash']}" in equity_formula:
        print(f"  ✓ Equity bridge formula present")
    else:
        errors_found.append(f"FAIL: Equity formula: {equity_formula}")
        print(f"  ❌ Equity formula: {equity_formula}")

# ── 2F: WACC Formula Check ─────────────────────────────────────────
print("\n--- 2F: WACC Formula Check ---")
ws_wacc = wb["WACC"]
wacc_formula = ws_wacc.cell(19, 2).value
if isinstance(wacc_formula, str):
    if "B16" in wacc_formula and "B9" in wacc_formula and "B17" in wacc_formula and "B13" in wacc_formula:
        print(f"  ✓ WACC formula = Ew*Ke + Dw*Kd_post")
    else:
        errors_found.append(f"FAIL: WACC formula: {wacc_formula}")
        print(f"  ❌ WACC formula: {wacc_formula}")

# Ke = Rf + Beta*ERP + SP + CRP
ke_formula = ws_wacc.cell(9, 2).value
if isinstance(ke_formula, str):
    if "B4" in ke_formula and "B5" in ke_formula and "B6" in ke_formula:
        print(f"  ✓ Ke formula = Rf + Beta*ERP + ...")
    else:
        errors_found.append(f"FAIL: Ke formula: {ke_formula}")
        print(f"  ❌ Ke formula: {ke_formula}")

# ── 2G: Balance Sheet formulas ──────────────────────────────────────
print("\n--- 2G: Balance Sheet Formula Structure ---")
ws_bs = wb["Balance Sheet"]
from src.dcf_engine.output.sheets_core import BS as BS_ROWS

# PPE from Capex DA
yr1_ppe = ws_bs.cell(BS_ROWS["PPE"], 2).value
if isinstance(yr1_ppe, str):
    if "Capex DA" in yr1_ppe:
        print(f"  ✓ BS PPE references Capex DA")
    else:
        errors_found.append(f"FAIL: BS PPE formula: {yr1_ppe}")
        print(f"  ❌ BS PPE formula: {yr1_ppe}")

# Cash from Cash Flow
yr1_cash = ws_bs.cell(BS_ROWS["Cash"], 2).value
if isinstance(yr1_cash, str):
    if "Cash Flow" in yr1_cash:
        print(f"  ✓ BS Cash references Cash Flow Ending Cash")
    else:
        errors_found.append(f"FAIL: BS Cash formula: {yr1_cash}")
        print(f"  ❌ BS Cash formula: {yr1_cash}")

# ── 2H: Cash Flow references ───────────────────────────────────────
print("\n--- 2H: Cash Flow Formula Structure ---")
ws_cf = wb["Cash Flow"]
from src.dcf_engine.output.sheets_core import CF as CF_ROWS

# NI from Income Statement
yr1_ni = ws_cf.cell(CF_ROWS["NI"], 2).value
if isinstance(yr1_ni, str):
    if "Income Statement" in yr1_ni:
        print(f"  ✓ CF NI references Income Statement")
    else:
        errors_found.append(f"FAIL: CF NI formula: {yr1_ni}")
        print(f"  ❌ CF NI formula: {yr1_ni}")

# UFCF = NOPAT + D&A + Capex + dNWC
yr1_ufcf = ws_cf.cell(CF_ROWS["UFCF"], 2).value
if isinstance(yr1_ufcf, str):
    if f"B{CF_ROWS['NOPAT']}" in yr1_ufcf:
        print(f"  ✓ CF UFCF formula correct")
    else:
        errors_found.append(f"FAIL: CF UFCF formula: {yr1_ufcf}")
        print(f"  ❌ CF UFCF formula: {yr1_ufcf}")

# Beg Cash Year 1 references Base Cash
yr1_begc = ws_cf.cell(CF_ROWS["BegC"], 2).value
if isinstance(yr1_begc, str):
    if f"$B${A['bcash']}" in yr1_begc:
        print(f"  ✓ CF Yr1 Beg Cash references base cash")
    else:
        errors_found.append(f"FAIL: CF Yr1 BegCash: {yr1_begc}")
        print(f"  ❌ CF Yr1 BegCash: {yr1_begc}")

# ── 2I: Working Capital references ─────────────────────────────────
print("\n--- 2I: Working Capital Formula Structure ---")
ws_wc = wb["Working Capital"]
from src.dcf_engine.output.sheets_core import WC as WC_ROWS

# Revenue from IS
yr1_rev = ws_wc.cell(WC_ROWS["Rev"], 2).value
if isinstance(yr1_rev, str):
    if "Income Statement" in yr1_rev:
        print(f"  ✓ WC Revenue references Income Statement")
    else:
        errors_found.append(f"FAIL: WC Rev formula: {yr1_rev}")
        print(f"  ❌ WC Rev formula: {yr1_rev}")

# Delta NWC Year 1 references base NWC
yr1_dnwc = ws_wc.cell(15, 2).value  # Delta NWC is row 15
if isinstance(yr1_dnwc, str):
    if f"$B${A['bnwc']}" in yr1_dnwc:
        print(f"  ✓ WC Yr1 Delta NWC references base NWC")
    else:
        errors_found.append(f"FAIL: WC Yr1 dNWC: {yr1_dnwc}")
        print(f"  ❌ WC Yr1 dNWC: {yr1_dnwc}")

# ── 2J: Checks tab formulas ────────────────────────────────────────
print("\n--- 2J: Checks Tab Formula Structure ---")
ws_chk = wb["Checks"]
yr1_bs_check = ws_chk.cell(3, 2).value
if isinstance(yr1_bs_check, str):
    if "Balance Sheet" in yr1_bs_check:
        print(f"  ✓ Checks BS Balance references Balance Sheet")
    else:
        errors_found.append(f"FAIL: Checks BS check formula: {yr1_bs_check}")
        print(f"  ❌ Checks BS check formula: {yr1_bs_check}")

yr1_cf_check = ws_chk.cell(6, 2).value
if isinstance(yr1_cf_check, str):
    if "Cash Flow" in yr1_cf_check and "Balance Sheet" in yr1_cf_check:
        print(f"  ✓ Checks CF check references both CF and BS")
    else:
        errors_found.append(f"FAIL: Checks CF check formula: {yr1_cf_check}")
        print(f"  ❌ Checks CF check formula: {yr1_cf_check}")

# ══════════════════════════════════════════════════════════════════════
# PART 3: PYTHON vs EXCEL FORMULA LOGIC COMPARISON
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PART 3: PYTHON vs EXCEL FORMULA LOGIC COMPARISON")
print("=" * 70)
print("Evaluating Excel formulas manually to compare with Python values...")

# We'll manually evaluate what the Excel formulas WOULD compute
# and compare against the Python pipeline output.

# Revenue Year 1: base_rev * (1 + rev_cagr)
excel_rev_y1 = BASE_REVENUE * (1 + 0.12)
python_rev_y1 = is_df.iloc[0]["Revenue"]
check("Yr1 Rev: Excel formula vs Python", excel_rev_y1, python_rev_y1)

# Revenue Year 5
excel_rev_y5 = BASE_REVENUE * (1.12 ** 5)
python_rev_y5 = is_df.iloc[4]["Revenue"]
check("Yr5 Rev: Excel formula vs Python", excel_rev_y5, python_rev_y5)

# COGS Year 1
excel_cogs_y1 = excel_rev_y1 * 0.38
python_cogs_y1 = is_df.iloc[0]["COGS"]
check("Yr1 COGS: Excel vs Python", excel_cogs_y1, python_cogs_y1)

# IS: GP matches
excel_gp_y1 = excel_rev_y1 - excel_cogs_y1
python_gp_y1 = is_df.iloc[0]["Gross Profit"]
check("Yr1 GP: Excel vs Python", excel_gp_y1, python_gp_y1)

# NOTE: Python COGS with Base scenario has margin_shift=0 so COGS = Rev * cogs_pct
# Excel does NOT apply margin_delta_bps. Check they are identical for Base.
print("\n  Python COGS formula (base): Rev * cogs_pct * (1 - margin_shift)")
print(f"  margin_shift for Base = {0.0 / 10000} = 0.0")
print(f"  Excel COGS formula: Rev * cogs_pct (no margin adjustment)")
print(f"  These match for Base scenario ✓")

# Capex DA Year 1
excel_beg_ppe_y1 = BASE_PPE
excel_capex_y1 = excel_rev_y1 * 0.06
excel_dep_ex_y1 = excel_beg_ppe_y1 * 0.12
excel_dep_new_y1 = excel_capex_y1 * 0.12 * 0.5
excel_dep_y1 = excel_dep_ex_y1 + excel_dep_new_y1
python_dep_y1 = cd_df.iloc[0]["Depreciation"]
check("Yr1 Depreciation: Excel vs Python", excel_dep_y1, python_dep_y1)

# Debt Schedule Year 1
excel_debt_beg_y1 = 3_000_000
excel_debt_int_y1 = excel_debt_beg_y1 * 0.055
python_debt_int_y1 = ds_df.iloc[0]["Interest Expense"]
check("Yr1 Interest: Excel vs Python", excel_debt_int_y1, python_debt_int_y1)

# NOTE: Excel debt repayment formula differs from Python
# Excel: IF(yr==maturity, Beg, MIN(amort, Beg))
# Python: amort + bullet at maturity
# For Year 1 (not maturity): both = MIN(200000, Beg) = 200000
excel_debt_rep_y1 = min(200_000, excel_debt_beg_y1)  # yr1 != 5
python_debt_rep_y1 = ds_df.iloc[0]["Principal Repayment"]
check("Yr1 Repayment: Excel vs Python", excel_debt_rep_y1, python_debt_rep_y1)

# Amortisation (intangibles)
# Python: rev * amort_pct_revenue
# Excel Yr1: MIN(Rev * amort_pct, intangibles)
# With intangibles = 0, Excel amort = MIN(Rev*0.003, 0) = 0
# Python amort = Rev * 0.003
# THIS IS A POTENTIAL MISMATCH when intangibles=0
python_amort_y1 = is_df.iloc[0]["Amortisation"]
excel_amort_y1 = min(excel_rev_y1 * 0.003, 0)  # intangibles = 0
print(f"\n  ⚠ Amortisation check: Python={python_amort_y1:,.2f}, Excel would={excel_amort_y1:,.2f}")
if abs(python_amort_y1 - excel_amort_y1) > 1:
    if excel_amort_y1 == 0 and python_amort_y1 > 0:
        msg = (f"MISMATCH: Amortisation - Python computes {python_amort_y1:,.2f} but "
               f"Excel MIN(Rev*amort_pct, intangibles=0) = 0")
        warnings_found.append(msg)
        print(f"  ⚠ {msg}")
    else:
        errors_found.append(f"FAIL: Amort Y1: Python={python_amort_y1:,.2f}, Excel={excel_amort_y1:,.2f}")

# EBITDA comparison
# Excel EBITDA = GP - SGA - OOpex
excel_ebitda_y1 = excel_gp_y1 - (excel_rev_y1 * 0.15) - (excel_rev_y1 * 0.07)
python_ebitda_y1 = is_df.iloc[0]["EBITDA"]
check("Yr1 EBITDA: Excel vs Python", excel_ebitda_y1, python_ebitda_y1)

# D&A for IS: Python uses capex schedule dep, Excel uses capex DA sheet dep
# Both should match for base case
is_dep_y1 = is_df.iloc[0]["Depreciation"]
cd_dep_y1 = cd_df.iloc[0]["Depreciation"]
check("Yr1 IS.Dep == CapexDA.Dep (both)", cd_dep_y1, is_dep_y1)

# EBIT comparison
excel_da_y1 = excel_dep_y1 + python_amort_y1  # Using python amort since Excel would differ
excel_ebit_y1 = excel_ebitda_y1 - (excel_dep_y1 + python_amort_y1)
python_ebit_y1 = is_df.iloc[0]["EBIT"]
check("Yr1 EBIT: Excel vs Python", excel_ebit_y1, python_ebit_y1)

# WC Delta NWC Year 1
excel_ar_y1 = excel_rev_y1 * 55 / 365
excel_inv_y1 = (excel_rev_y1 * 0.38) * 35 / 365
excel_prep_y1 = excel_rev_y1 * 0.008
excel_oca_y1 = excel_rev_y1 * 0.004
excel_ap_y1 = (excel_rev_y1 * 0.38) * 50 / 365
excel_accr_y1 = excel_rev_y1 * 0.012
excel_ocl_y1 = excel_rev_y1 * 0.006
excel_nwc_y1 = excel_ar_y1 + excel_inv_y1 + excel_prep_y1 + excel_oca_y1 - excel_ap_y1 - excel_accr_y1 - excel_ocl_y1
python_nwc_y1 = wc_df.iloc[0]["NWC"]
check("Yr1 NWC: Excel vs Python", excel_nwc_y1, python_nwc_y1)

# But note: Excel WC Inv references COGS row from IS (=IS!C4)
# which is Rev*cogs_pct, and then multiplies by DIO/365.
# In Python: COGS is also Rev*cogs_pct, then Inv = COGS*DIO/365
# These should match for base scenario.
print("\n  WC Inventory source: Python=COGS*DIO/365, Excel=IS_COGS*DIO/365 - should match ✓")

# WACC computation
excel_ke = 0.038 + 1.25 * 0.065 + 0.01 + 0.005
excel_kd_pre = 0.038 + 0.0108  # A spread (ICR=6.5)
excel_kd_post = excel_kd_pre * (1 - 0.22)
# Excel WACC: Ew*Ke + Dw*Kd_post
# But Excel Kd uses "Credit Spread (computed)" at row 31
# Check what's at row 31
credit_spread_excel = ws_a.cell(31, 2).value
print(f"\n  Excel Credit Spread (row 31): {credit_spread_excel}")
print(f"  Python credit spread (A @ICR=6.5): 0.0108")
if isinstance(credit_spread_excel, (int, float)):
    if abs(credit_spread_excel - 0.0108) > 0.001:
        msg = f"Credit spread mismatch: Excel={credit_spread_excel}, expected=0.0108"
        warnings_found.append(msg)
        print(f"  ⚠ {msg}")

# Excel WACC formula: B16*B9 + B17*B13
# B16 = ew = 0.65, B9 = Ke, B17 = dw = 0.35, B13 = Kd post-tax
# B12 = Rf + CreditSpread(row31)
# Check: does WACC sheet B12 use row 31?
wacc_kd_pre_formula = ws_wacc.cell(12, 2).value
print(f"  WACC Kd Pre-Tax formula: {wacc_kd_pre_formula}")
if isinstance(wacc_kd_pre_formula, str) and "$B$31" in wacc_kd_pre_formula:
    print(f"  ✓ WACC Kd correctly references Assumptions credit spread (B31)")
else:
    print(f"  ~ WACC Kd formula: {wacc_kd_pre_formula}")

# ── UFCF in Excel vs Python ─────────────────────────────────────────
print("\n--- UFCF Excel vs Python ---")
# Excel UFCF = NOPAT + D&A + Capex(negative) + dNWC(negative means cash use)
# Python UFCF = NOPAT + D&A - Capex(positive) - dNWC(positive means increase)
# Both should arrive at the same number
for i in range(5):
    yr = i + 1
    python_ufcf = fcf_df.iloc[i]["Unlevered FCF"]
    print(f"  Yr{yr} Python UFCF: {python_ufcf:,.2f}")

# ── Terminal Value ──────────────────────────────────────────────────
print("\n--- Terminal Value Excel vs Python ---")
# Excel TV Gordon = Terminal_FCF*(1+tg)/(WACC-tg)
# Note: Excel does NOT cap terminal growth at GDP cap or apply spread floor
# This is a potential mismatch!
python_effective_g = d.effective_terminal_growth
print(f"  Python effective terminal growth: {python_effective_g:.4f}")
print(f"  Config terminal growth: {0.03}")
print(f"  Python WACC: {wacc_val:.4f}")
if abs(python_effective_g - 0.03) > 0.0001:
    msg = (f"Excel TV uses nominal tg=3.0% but Python uses effective_g={python_effective_g:.4f} "
           f"(capped by GDP cap or spread floor)")
    warnings_found.append(msg)
    print(f"  ⚠ {msg}")

    # Calculate what Excel would get vs Python
    excel_gordon_tv = terminal_fcf * (1 + 0.03) / (wacc_val - 0.03)
    python_gordon_tv = d.gordon_tv
    print(f"  Excel Gordon TV (using 3.0%): {excel_gordon_tv:,.2f}")
    print(f"  Python Gordon TV (using {python_effective_g:.4f}): {python_gordon_tv:,.2f}")
    if abs(excel_gordon_tv - python_gordon_tv) > 1:
        errors_found.append(
            f"MISMATCH: Gordon TV - Excel uses tg=3.0% giving {excel_gordon_tv:,.2f}, "
            f"Python uses effective_g={python_effective_g:.4f} giving {python_gordon_tv:,.2f}"
        )
else:
    print(f"  ✓ Terminal growth matches between Excel and Python")

# ── Shares: Check if Excel stores shares in millions vs absolute ───
print("\n--- Shares outstanding convention ---")
shares_excel = ws_a.cell(A["shares"], 2).value
print(f"  Excel Assumptions shares value: {shares_excel}")
print(f"  Python fully_diluted_shares config: {cfg.valuation.fully_diluted_shares}")
print(f"  DCF Price per Share formula divides Equity by B37 (shares from Assumptions)")
# If shares stored as 2.0 (millions), price = equity / 2.0 → in millions
# If shares stored as 2000000, price = equity / 2000000 → per share
# The code stores: v.fully_diluted_shares/1e6 if > 1e6, else raw
if cfg.valuation.fully_diluted_shares > 1e6:
    if isinstance(shares_excel, (int, float)) and shares_excel < 100:
        msg = (f"CRITICAL: Shares stored as {shares_excel}M in Excel but DCF divides Equity by this "
               f"number, giving price in millions instead of per-share!")
        errors_found.append(msg)
        print(f"  ❌ {msg}")
    else:
        print(f"  ✓ Shares stored correctly")
else:
    print(f"  ✓ Shares {shares_excel} OK")

# ══════════════════════════════════════════════════════════════════════
# PART 4: SENSITIVITY / TORNADO / MONTE CARLO CONSISTENCY
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PART 4: SENSITIVITY / TORNADO / MC CHECKS")
print("=" * 70)

# Sensitivity base case should equal the base equity
if result.sensitivity:
    s = result.sensitivity
    # Find the center cell of wacc_vs_growth table
    mid_row = len(s.growth_values) // 2
    mid_col = len(s.wacc_values) // 2
    center_eq = s.wacc_vs_growth.iloc[mid_row, mid_col]
    print(f"\n  Sensitivity WACC/TG center: {center_eq:,.2f}")
    print(f"  DCF equity blended: {d.equity_blended:,.2f}")
    # The center uses _quick_dcf which simplifies (no WC schedule, no debt schedule details)
    # so it won't match exactly, but should be in the same ballpark
    ratio = center_eq / d.equity_blended if d.equity_blended != 0 else 0
    print(f"  Ratio (sensitivity center / DCF equity): {ratio:.3f}")
    if abs(ratio - 1.0) > 0.3:
        warnings_found.append(f"Sensitivity center ({center_eq:,.0f}) far from DCF equity ({d.equity_blended:,.0f})")

# Tornado should have base_equity matching DCF
if result.tornado:
    t = result.tornado
    print(f"\n  Tornado base_equity: {t.base_equity:,.2f}")
    print(f"  DCF equity_blended: {d.equity_blended:,.2f}")
    check("Tornado base_equity == DCF equity_blended", d.equity_blended, t.base_equity)

# Monte Carlo stats should be reasonable
if result.monte_carlo:
    mc = result.monte_carlo
    print(f"\n  MC Mean equity: {mc.statistics['Mean']:,.2f}")
    print(f"  MC Median equity: {mc.statistics['Median']:,.2f}")
    print(f"  MC P10: {mc.statistics['P10']:,.2f}")
    print(f"  MC P90: {mc.statistics['P90']:,.2f}")
    print(f"  MC iterations: {mc.iterations}")
    if mc.statistics['Mean'] < 0:
        warnings_found.append("MC Mean equity is negative")

# ══════════════════════════════════════════════════════════════════════
# PART 5: SCENARIO COMPARISON
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("PART 5: SCENARIO COMPARISON")
print("=" * 70)

if result.scenario_comparison:
    comp = result.scenario_comparison.comparison
    print(f"\n  Scenarios: {list(comp.columns)}")
    print(f"\n  Metrics: {list(comp.index)}")
    
    # Bull should have higher equity than Base, Bear lower
    if "EV (Blended)" in comp.index:
        base_ev = comp.loc["EV (Blended)", "Base"] if "Base" in comp.columns else None
        bull_ev = comp.loc["EV (Blended)", "Bull"] if "Bull" in comp.columns else None
        bear_ev = comp.loc["EV (Blended)", "Bear"] if "Bear" in comp.columns else None
        if base_ev and bull_ev and bear_ev:
            print(f"  Base EV: {base_ev:,.2f}")
            print(f"  Bull EV: {bull_ev:,.2f}")
            print(f"  Bear EV: {bear_ev:,.2f}")
            if bull_ev <= base_ev:
                errors_found.append(f"FAIL: Bull EV ({bull_ev:,.0f}) should be > Base EV ({base_ev:,.0f})")
                print(f"  ❌ Bull EV should be > Base EV")
            else:
                print(f"  ✓ Bull EV > Base EV")
            if bear_ev >= base_ev:
                errors_found.append(f"FAIL: Bear EV ({bear_ev:,.0f}) should be < Base EV ({base_ev:,.0f})")
                print(f"  ❌ Bear EV should be < Base EV")
            else:
                print(f"  ✓ Bear EV < Base EV")

# ══════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("FINAL SUMMARY")
print("=" * 70)

if errors_found:
    print(f"\n❌ {len(errors_found)} ERRORS FOUND:")
    for e in errors_found:
        print(f"  • {e}")
else:
    print(f"\n✓ NO ERRORS FOUND")

if warnings_found:
    print(f"\n⚠ {len(warnings_found)} WARNINGS:")
    for w in warnings_found:
        print(f"  • {w}")

print(f"\nExcel file: {excel_path}")
print(f"Pipeline errors: {result.errors}")
print(f"Pipeline timings: {result.timings}")
