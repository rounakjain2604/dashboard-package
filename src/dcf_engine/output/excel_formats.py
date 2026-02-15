"""Shared Excel formatting constants and helpers for the IB-grade workbook."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

# ── Fonts (Verdana throughout) ───────────────────────────────────────
FH = Font(name="Verdana", size=9, bold=True, color="FFFFFF")
FS = Font(name="Verdana", size=9, bold=True, color="1F3864")
FN = Font(name="Verdana", size=9)
FI = Font(name="Verdana", size=9, color="0000CC")
FT = Font(name="Verdana", size=12, bold=True, color="1F3864")
FB = Font(name="Verdana", size=9, bold=True)
FK = Font(name="Verdana", size=11, bold=True, color="1F3864")

# ── Fills ────────────────────────────────────────────────────────────
FILL_HDR = PatternFill("solid", fgColor="1F3864")
FILL_INP = PatternFill("solid", fgColor="DCE6F1")
FILL_TOT = PatternFill("solid", fgColor="E2EFDA")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

# ── Alignment & Borders ─────────────────────────────────────────────
AC = Alignment(horizontal="center", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AL = Alignment(horizontal="left", vertical="center")
BORD = Border(bottom=Side(style="thin", color="D9D9D9"))
BORD_M = Border(bottom=Side(style="medium", color="1F3864"))

# ── Number Formats ───────────────────────────────────────────────────
NF = '#,##0'
PF = '0.0%'
XF = '0.0x'
PF2 = '$#,##0.00'

# ── Assumptions sheet row map (column B) ─────────────────────────────
A = {
    "rev_cagr": 4, "cogs_pct": 5, "sga_pct": 6, "oopex_pct": 7,
    "dep_rate": 8, "capex_pct": 9, "tax": 10,
    "dso": 11, "dio": 12, "dpo": 13, "div": 14,
    "amort_pct": 15, "prepaid_pct": 16, "accrued_pct": 17,
    "oca_pct": 18, "ocl_pct": 19,
    "rf": 23, "erp": 24, "beta": 25, "sp": 26, "crp": 27,
    "dw": 28, "ew": 29, "icr": 30,
    "tg": 34, "exit_m": 35, "gw": 36, "shares": 37, "cash": 38, "debt": 39,
    "brev": 43, "bcash": 44, "bppe": 45, "bnwc": 46, "bre": 47, "bcs": 48,
    # Debt Schedule Inputs
    "debt_bal": 51, "debt_rate": 52, "debt_amort": 53, "debt_maturity": 54,
    # Other Balance Sheet Items
    "goodwill": 57, "intangibles": 58, "olt_assets": 59, "olt_liab": 60,
    # Capex Method (for fixed vs pct_revenue)
    "capex_method": 62, "capex_fixed": 63,
}

def ar(key):
    """Absolute reference to Assumptions!$B$<row>."""
    return f"Assumptions!$B${A[key]}"


def sr(sheet, c, r):
    """Reference to cell on another sheet."""
    if " " in sheet:
        return f"'{sheet}'!{c}{r}"
    return f"{sheet}!{c}{r}"


def hdr_row(ws, row, ncols):
    """Style a header row."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FH
        cell.fill = FILL_HDR
        cell.alignment = AC


def style_label(ws, row, col, label, bold=False, section=False):
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = FS if section else (FB if bold else FN)
    if bold:
        cell.fill = FILL_TOT
        cell.border = BORD_M
    return cell


def fmt_cells(ws, row, start_col, end_col, num_format=NF, bold=False):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.number_format = num_format
        cell.font = FB if bold else FN
        cell.alignment = AR
        cell.border = BORD
        if bold:
            cell.fill = FILL_TOT
            cell.border = BORD_M


def fit_to_width(ws):
    """Set print to fit page width."""
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def auto_col(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


CL = get_column_letter
