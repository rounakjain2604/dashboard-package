"""Formula-linked sheets: Balance Sheet, Cash Flow, WACC, DCF."""
from __future__ import annotations
from openpyxl.utils import get_column_letter as CL
from .excel_formats import *
from .sheets_core import IS, WC, CD, DS, BS, CF, WR, DC


def build_bs(wb, n):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_properties.tabColor = "548235"
    auto_col(ws,1,24)
    ws.cell(1,1,value="Balance Sheet").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    labels = {
        BS["Cash"]:"Cash & Equivalents", BS["AR"]:"Accounts Receivable",
        BS["Inv"]:"Inventory", BS["Prep"]:"Prepaid Expenses",
        BS["OCA"]:"Other Current Assets", BS["CA"]:"Total Current Assets",
        BS["PPE"]:"PP&E (Net)", BS["GW"]:"Goodwill", BS["Intan"]:"Intangibles",
        BS["OLA"]:"Other LT Assets", BS["NCA"]:"Total Non-Current Assets",
        BS["TA"]:"TOTAL ASSETS",
        14:"", BS["AP"]:"Accounts Payable", BS["Accr"]:"Accrued Liabilities",
        BS["OCL"]:"Other Current Liabilities", BS["CurD"]:"Current Portion of Debt",
        BS["CL"]:"Total Current Liabilities",
        BS["LTD"]:"Long-Term Debt", BS["OLL"]:"Other LT Liabilities",
        BS["NCL"]:"Total Non-Current Liabilities", BS["TL"]:"TOTAL LIABILITIES",
        24:"", BS["CS"]:"Common Stock", BS["RE"]:"Retained Earnings",
        BS["TE"]:"Total Equity", BS["TLE"]:"TOTAL L + E",
        29:"", BS["Chk"]:"Balance Check (A-L-E)", BS["Stat"]:"Status",
    }
    bolds = {BS["CA"],BS["NCA"],BS["TA"],BS["CL"],BS["NCL"],BS["TL"],BS["TE"],BS["TLE"],BS["Chk"]}
    for r,lbl in labels.items():
        if lbl: style_label(ws,r,1,lbl,bold=(r in bolds))
    for j in range(n):
        col=j+2; c=CL(col); pc=CL(col-1) if j>0 else None
        ws[f"{c}{BS['Cash']}"]  = f"='Cash Flow'!{c}{CF['EndC']}"
        ws[f"{c}{BS['AR']}"]    = f"='Working Capital'!{c}{WC['AR']}"
        ws[f"{c}{BS['Inv']}"]   = f"='Working Capital'!{c}{WC['Inv']}"
        ws[f"{c}{BS['Prep']}"]  = f"='Working Capital'!{c}{WC['Prep']}"
        ws[f"{c}{BS['OCA']}"]   = f"='Working Capital'!{c}{WC['OCA']}"
        ws[f"{c}{BS['CA']}"]    = f"=SUM({c}{BS['Cash']}:{c}{BS['OCA']})"
        ws[f"{c}{BS['PPE']}"]   = f"='Capex DA'!{c}{CD['End']}"
        ws[f"{c}{BS['GW']}"]    = f"={ar('goodwill')}"
        if j==0:
            ws[f"{c}{BS['Intan']}"] = f"=MAX({ar('intangibles')}-'Income Statement'!{c}{IS['Amort']},0)"
        else:
            ws[f"{c}{BS['Intan']}"] = f"=MAX({pc}{BS['Intan']}-'Income Statement'!{c}{IS['Amort']},0)"
        ws[f"{c}{BS['OLA']}"]   = f"={ar('olt_assets')}"
        ws[f"{c}{BS['NCA']}"]   = f"=SUM({c}{BS['PPE']}:{c}{BS['OLA']})"
        ws[f"{c}{BS['TA']}"]    = f"={c}{BS['CA']}+{c}{BS['NCA']}"
        ws[f"{c}{BS['AP']}"]    = f"='Working Capital'!{c}{WC['AP']}"
        ws[f"{c}{BS['Accr']}"]  = f"='Working Capital'!{c}{WC['Accr']}"
        ws[f"{c}{BS['OCL']}"]   = f"='Working Capital'!{c}{WC['OCL']}"
        ws[f"{c}{BS['CurD']}"]  = f"='Debt Schedule'!{c}{DS['Cur']}"
        ws[f"{c}{BS['CL']}"]    = f"=SUM({c}{BS['AP']}:{c}{BS['CurD']})"
        ws[f"{c}{BS['LTD']}"]   = f"='Debt Schedule'!{c}{DS['End']}-{c}{BS['CurD']}"
        ws[f"{c}{BS['OLL']}"]   = f"={ar('olt_liab')}"
        ws[f"{c}{BS['NCL']}"]   = f"={c}{BS['LTD']}+{c}{BS['OLL']}"
        ws[f"{c}{BS['TL']}"]    = f"={c}{BS['CL']}+{c}{BS['NCL']}"
        ws[f"{c}{BS['CS']}"]    = f"={ar('bcs')}"
        if j==0:
            ws[f"{c}{BS['RE']}"] = f"={ar('bre')}+'Income Statement'!{c}{IS['NI']}-'Income Statement'!{c}{IS['NI']}*{ar('div')}"
        else:
            ws[f"{c}{BS['RE']}"] = f"={pc}{BS['RE']}+'Income Statement'!{c}{IS['NI']}-'Income Statement'!{c}{IS['NI']}*{ar('div')}"
        ws[f"{c}{BS['TE']}"]    = f"={c}{BS['CS']}+{c}{BS['RE']}"
        ws[f"{c}{BS['TLE']}"]   = f"={c}{BS['TL']}+{c}{BS['TE']}"
        ws[f"{c}{BS['Chk']}"]   = f"={c}{BS['TA']}-{c}{BS['TLE']}"
        ws[f"{c}{BS['Stat']}"]  = f'=IF(ABS({c}{BS["Chk"]})<1,"PASS","FAIL")'
        for r in range(2,32):
            if r in {14,24,29}: continue
            fmt_cells(ws,r,col,col,NF,bold=(r in bolds))
    fit_to_width(ws)


def build_cf(wb, n):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_properties.tabColor = "4472C4"
    auto_col(ws,1,26)
    ws.cell(1,1,value="Cash Flow Statement").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    labels = {
        2:"CASH FROM OPERATIONS",
        CF["NI"]:"Net Income", CF["Dep"]:"Depreciation", CF["Amort"]:"Amortisation",
        CF["DA"]:"Total D&A", CF["dNWC"]:"Change in NWC", CF["CFO"]:"Cash from Operations",
        9:"", 10:"CASH FROM INVESTING",
        CF["Capex"]:"Capital Expenditures", CF["CFI"]:"Cash from Investing",
        13:"", 14:"CASH FROM FINANCING",
        CF["NewD"]:"New Debt Issuance", CF["RepD"]:"Debt Repayment",
        CF["IntP"]:"Interest Paid", CF["Div"]:"Dividends Paid",
        CF["CFF"]:"Cash from Financing",
        20:"", CF["Net"]:"Net Change in Cash",
        CF["BegC"]:"Beginning Cash", CF["EndC"]:"Ending Cash",
        24:"", 25:"UNLEVERED FREE CASH FLOW",
        CF["EBIT"]:"EBIT", CF["NOPAT"]:"NOPAT (after tax)",
        CF["DA2"]:"Add: D&A", CF["Capex2"]:"Less: Capex",
        CF["dNWC2"]:"Less: Change in NWC", CF["UFCF"]:"Unlevered FCF",
    }
    bolds = {CF["CFO"],CF["CFI"],CF["CFF"],CF["Net"],CF["EndC"],CF["UFCF"]}
    sections = {2,10,14,25}
    for r,lbl in labels.items():
        if lbl:
            if r in sections: style_label(ws,r,1,lbl,section=True)
            else: style_label(ws,r,1,lbl,bold=(r in bolds))
    for j in range(n):
        col=j+2; c=CL(col); pc=CL(col-1) if j>0 else None
        # CFO
        ws[f"{c}{CF['NI']}"]   = f"='Income Statement'!{c}{IS['NI']}"
        ws[f"{c}{CF['Dep']}"]  = f"='Capex DA'!{c}{CD['Dep']}"
        ws[f"{c}{CF['Amort']}"]= f"='Income Statement'!{c}{IS['Amort']}"
        ws[f"{c}{CF['DA']}"]   = f"={c}{CF['Dep']}+{c}{CF['Amort']}"
        ws[f"{c}{CF['dNWC']}"] = f"=-'Working Capital'!{c}{WC['dNWC']}"
        ws[f"{c}{CF['CFO']}"]  = f"={c}{CF['NI']}+{c}{CF['DA']}+{c}{CF['dNWC']}"
        # CFI
        ws[f"{c}{CF['Capex']}"]= f"=-'Capex DA'!{c}{CD['Capex']}"
        ws[f"{c}{CF['CFI']}"]  = f"={c}{CF['Capex']}"
        # CFF
        ws[f"{c}{CF['NewD']}"] = f"='Debt Schedule'!{c}{DS['New']}"
        ws[f"{c}{CF['RepD']}"] = f"=-'Debt Schedule'!{c}{DS['Rep']}"
        ws[f"{c}{CF['IntP']}"] = f"=0"
        ws[f"{c}{CF['Div']}"]  = f"=-'Income Statement'!{c}{IS['NI']}*{ar('div')}"
        ws[f"{c}{CF['CFF']}"]  = f"={c}{CF['NewD']}+{c}{CF['RepD']}+{c}{CF['Div']}"
        # Net / Cash
        ws[f"{c}{CF['Net']}"]  = f"={c}{CF['CFO']}+{c}{CF['CFI']}+{c}{CF['CFF']}"
        if j==0:
            ws[f"{c}{CF['BegC']}"] = f"={ar('bcash')}"
        else:
            ws[f"{c}{CF['BegC']}"] = f"={pc}{CF['EndC']}"
        ws[f"{c}{CF['EndC']}"] = f"={c}{CF['BegC']}+{c}{CF['Net']}"
        # UFCF
        ws[f"{c}{CF['EBIT']}"]  = f"='Income Statement'!{c}{IS['EBIT']}"
        ws[f"{c}{CF['NOPAT']}"] = f"={c}{CF['EBIT']}*(1-{ar('tax')})"
        ws[f"{c}{CF['DA2']}"]   = f"={c}{CF['DA']}"
        ws[f"{c}{CF['Capex2']}"]= f"={c}{CF['Capex']}"
        ws[f"{c}{CF['dNWC2']}"] = f"={c}{CF['dNWC']}"
        ws[f"{c}{CF['UFCF']}"]  = f"={c}{CF['NOPAT']}+{c}{CF['DA2']}+{c}{CF['Capex2']}+{c}{CF['dNWC2']}"
        for r in range(2,32):
            if r in {9,13,20,24,2,10,14,25}: continue
            fmt_cells(ws,r,col,col,NF,bold=(r in bolds))
    fit_to_width(ws)


def build_wacc(wb, credit_spread=0.02):
    ws = wb.create_sheet("WACC")
    ws.sheet_properties.tabColor = "7030A0"
    auto_col(ws,1,28); auto_col(ws,2,18)
    ws.cell(1,1,value="WACC Calculation").font = FK
    ws.merge_cells("A1:B1")
    style_label(ws,3,1,"COST OF EQUITY (CAPM)",section=True)
    for r,lbl in {4:"Risk-Free Rate",5:"Beta",6:"Equity Risk Premium",7:"Size Premium",8:"Country Risk Premium",9:"Cost of Equity (Ke)"}.items():
        style_label(ws,r,1,lbl,bold=(r==9))
    ws["B4"] = f"={ar('rf')}";      ws["B4"].number_format=PF
    ws["B5"] = f"={ar('beta')}";    ws["B5"].number_format='0.00'
    ws["B6"] = f"={ar('erp')}";     ws["B6"].number_format=PF
    ws["B7"] = f"={ar('sp')}";      ws["B7"].number_format=PF
    ws["B8"] = f"={ar('crp')}";     ws["B8"].number_format=PF
    ws["B9"] = f"=B4+B5*B6+B7+B8";  ws["B9"].number_format=PF; ws["B9"].font=FB
    style_label(ws,11,1,"COST OF DEBT",section=True)
    for r,lbl in {12:"Pre-Tax Kd",13:"After-Tax Kd"}.items():
        style_label(ws,r,1,lbl,bold=(r==13))
    ws["B12"] = f"={ar('rf')}+Assumptions!$B$31"; ws["B12"].number_format=PF
    ws["B13"] = f"=B12*(1-{ar('tax')})"; ws["B13"].number_format=PF; ws["B13"].font=FB
    style_label(ws,15,1,"CAPITAL STRUCTURE",section=True)
    for r,lbl in {16:"Equity Weight",17:"Debt Weight"}.items():
        style_label(ws,r,1,lbl)
    ws["B16"] = f"={ar('ew')}"; ws["B16"].number_format=PF
    ws["B17"] = f"={ar('dw')}"; ws["B17"].number_format=PF
    style_label(ws,19,1,"WACC",bold=True)
    ws["B19"] = f"=B16*B9+B17*B13"; ws["B19"].number_format=PF; ws["B19"].font=FB
    ws["B19"].fill = FILL_TOT
    fit_to_width(ws)


def build_dcf(wb, n):
    ws = wb.create_sheet("DCF")
    ws.sheet_properties.tabColor = "C00000"
    auto_col(ws,1,26)
    ws.cell(1,1,value="DCF Valuation").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    for r,lbl in {DC["UFCF"]:"Unlevered FCF",DC["DF"]:"Discount Factor",DC["PV"]:"PV of UFCF"}.items():
        style_label(ws,r,1,lbl,bold=(r==DC["PV"]))
    lc = CL(n+1)  # last year column
    for j in range(n):
        col=j+2; c=CL(col)
        ws[f"{c}{DC['UFCF']}"] = f"='Cash Flow'!{c}{CF['UFCF']}"
        ws[f"{c}{DC['DF']}"]   = f"=1/(1+WACC!$B$19)^({j+1}-0.5)"
        ws[f"{c}{DC['PV']}"]   = f"={c}{DC['UFCF']}*{c}{DC['DF']}"
        for r in [DC["UFCF"],DC["PV"]]:
            fmt_cells(ws,r,col,col,NF,bold=(r==DC["PV"]))
        fmt_cells(ws,DC["DF"],col,col,'0.0000')
    # Sum PV
    style_label(ws,DC["SumPV"],1,"Sum of PV (FCFs)",bold=True)
    ws.cell(DC["SumPV"],2).value = f"=SUM(B{DC['PV']}:{lc}{DC['PV']})"
    ws.cell(DC["SumPV"],2).number_format=NF; ws.cell(DC["SumPV"],2).font=FB
    # Terminal Value section
    style_label(ws,8,1,"TERMINAL VALUE",section=True)
    for r,lbl in {DC["TFCF"]:"Terminal Year FCF",DC["TEBITDA"]:"Terminal Year EBITDA",
                   DC["TVG"]:"TV (Gordon Growth)",DC["TVE"]:"TV (Exit Multiple)",
                   DC["TVB"]:"TV (Blended)",DC["PVTV"]:"PV of Terminal Value"}.items():
        style_label(ws,r,1,lbl,bold=(r in {DC["TVB"],DC["PVTV"]}))
    ws.cell(DC["TFCF"],2).value = f"={lc}{DC['UFCF']}"; ws.cell(DC["TFCF"],2).number_format=NF
    ws.cell(DC["TEBITDA"],2).value = f"='Income Statement'!{lc}{IS['EBITDA']}"; ws.cell(DC["TEBITDA"],2).number_format=NF
    ws.cell(DC["TVG"],2).value = f"=B{DC['TFCF']}*(1+{ar('tg')})/(WACC!$B$19-{ar('tg')})"; ws.cell(DC["TVG"],2).number_format=NF
    ws.cell(DC["TVE"],2).value = f"=B{DC['TEBITDA']}*{ar('exit_m')}"; ws.cell(DC["TVE"],2).number_format=NF
    ws.cell(DC["TVB"],2).value = f"=B{DC['TVG']}*{ar('gw')}+B{DC['TVE']}*(1-{ar('gw')})"; ws.cell(DC["TVB"],2).number_format=NF; ws.cell(DC["TVB"],2).font=FB
    ws.cell(DC["PVTV"],2).value = f"=B{DC['TVB']}*{lc}{DC['DF']}"; ws.cell(DC["PVTV"],2).number_format=NF; ws.cell(DC["PVTV"],2).font=FB
    # Equity Bridge
    style_label(ws,16,1,"EQUITY BRIDGE",section=True)
    for r,lbl in {DC["EV"]:"Enterprise Value",DC["Cash"]:"Plus: Cash",DC["Debt"]:"Less: Net Debt",
                   DC["Equity"]:"Equity Value",DC["Shares"]:"Shares Outstanding",DC["Price"]:"Price per Share"}.items():
        style_label(ws,r,1,lbl,bold=(r in {DC["Equity"],DC["Price"]}))
    ws.cell(DC["EV"],2).value = f"=B{DC['SumPV']}+B{DC['PVTV']}"; ws.cell(DC["EV"],2).number_format=NF
    ws.cell(DC["Cash"],2).value = f"={ar('cash')}"; ws.cell(DC["Cash"],2).number_format=NF
    ws.cell(DC["Debt"],2).value = f"={ar('debt')}"; ws.cell(DC["Debt"],2).number_format=NF
    ws.cell(DC["Equity"],2).value = f"=B{DC['EV']}+B{DC['Cash']}-B{DC['Debt']}"; ws.cell(DC["Equity"],2).number_format=NF; ws.cell(DC["Equity"],2).font=FB
    ws.cell(DC["Shares"],2).value = f"={ar('shares')}"; ws.cell(DC["Shares"],2).number_format=NF
    ws.cell(DC["Price"],2).value = f"=IF(B{DC['Shares']}=0,0,B{DC['Equity']}/B{DC['Shares']})"; ws.cell(DC["Price"],2).number_format=PF2; ws.cell(DC["Price"],2).font=FB
    ws.cell(DC["Price"],2).fill=FILL_TOT
    fit_to_width(ws)
