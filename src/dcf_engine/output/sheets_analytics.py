"""Value-based tabs (simulation results) + Checks + Audit Trail."""
from __future__ import annotations
import numpy as np
from datetime import datetime
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter as CL
from .excel_formats import *
from .sheets_core import BS, WR


def _as_dict(item):
    if hasattr(item, "__dataclass_fields__"):
        from dataclasses import asdict
        return asdict(item)
    if isinstance(item, dict):
        return item
    return {"value": str(item)}


def _write_table(ws, title, rows, columns):
    ws.cell(1, 1, value=title).font = FK
    for j, (key, label, width) in enumerate(columns, 1):
        ws.cell(3, j, value=label).font = FH
        ws.cell(3, j).fill = FILL_HDR
        auto_col(ws, j, width)
    if not rows:
        ws.cell(4, 1, value="No data supplied").font = FN
        fit_to_width(ws)
        return
    for r, raw in enumerate(rows, 4):
        item = _as_dict(raw)
        for j, (key, _label, _width) in enumerate(columns, 1):
            value = item.get(key)
            cell = ws.cell(r, j, value=value)
            cell.font = FN
            cell.number_format = NF if isinstance(value, (int, float)) else "@"
    fit_to_width(ws)


def build_source_map(wb, source_facts):
    ws = wb.create_sheet("Source Map")
    ws.sheet_properties.tabColor = "0C6B58"
    _write_table(ws, "SEC Source Map", source_facts or [], [
        ("account", "Account", 24),
        ("concept", "Concept", 34),
        ("value", "Value", 18),
        ("unit", "Unit", 12),
        ("form", "Form", 12),
        ("filed", "Filed", 14),
        ("period_end", "Period", 14),
        ("source_url", "Source URL", 42),
    ])


def build_warnings(wb, warnings):
    ws = wb.create_sheet("Warnings")
    ws.sheet_properties.tabColor = "C00000"
    rows = [{"severity": "warning", "message": str(w)} for w in (warnings or [])]
    _write_table(ws, "Model Warnings", rows, [
        ("severity", "Severity", 14),
        ("message", "Message", 80),
    ])


def build_filing_changes(wb, changes):
    ws = wb.create_sheet("Filing Changes")
    ws.sheet_properties.tabColor = "9A4F00"
    _write_table(ws, "Numeric Filing Changes", changes or [], [
        ("account", "Account", 24),
        ("category", "Category", 18),
        ("severity", "Severity", 12),
        ("latest_value", "Latest", 18),
        ("prior_value", "Prior", 18),
        ("percent_change", "% Change", 16),
        ("valuation_impact", "Valuation Impact", 54),
    ])


def build_valuation_impacts(wb, impacts):
    ws = wb.create_sheet("Valuation Impacts")
    ws.sheet_properties.tabColor = "7030A0"
    _write_table(ws, "Valuation Impacts", impacts or [], [
        ("severity", "Severity", 12),
        ("title", "Title", 28),
        ("detail", "Detail", 70),
        ("affected_assumptions", "Affected Assumptions", 34),
    ])


def build_scenarios(wb, scenario_comparison):
    """Write scenario comparison table (values from simulation)."""
    ws = wb.create_sheet("Scenarios")
    ws.sheet_properties.tabColor = "BF8F00"
    if scenario_comparison is None or scenario_comparison.comparison.empty:
        ws.cell(1,1,value="No scenario data").font = FN; return
    comp = scenario_comparison.comparison
    ws.cell(1,1,value="Scenario Comparison").font = FK
    auto_col(ws,1,24)
    cols = list(comp.columns)
    ws.cell(2,1,value="Metric").font = FH; ws.cell(2,1).fill = FILL_HDR
    for j,cn in enumerate(cols,2):
        ws.cell(2,j,value=cn).font = FH; ws.cell(2,j).fill = FILL_HDR
        auto_col(ws,j,16)
    for i,metric in enumerate(comp.index,3):
        ws.cell(i,1,value=metric).font = FN
        for j,cn in enumerate(cols,2):
            v = comp.loc[metric,cn]
            cell = ws.cell(i,j,value=v if not (isinstance(v,float) and np.isnan(v)) else 0)
            cell.font = FN; cell.alignment = AR
            cell.number_format = PF if isinstance(v,float) and abs(v)<2 else NF
    fit_to_width(ws)


def build_sensitivity(wb, sensitivity_result):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_properties.tabColor = "305496"
    if sensitivity_result is None:
        ws.cell(1,1,value="No sensitivity data").font = FN; return
    auto_col(ws,1,20)
    r = 1
    for title,tbl in [("WACC vs Terminal Growth",sensitivity_result.wacc_vs_growth),
                       ("Revenue Growth vs EBITDA Margin",sensitivity_result.revenue_vs_margin)]:
        ws.cell(r,1,value=title).font = FK; r += 1
        df = tbl.reset_index()
        for j,cn in enumerate(df.columns):
            ws.cell(r,j+1,value=str(cn)).font = FH; ws.cell(r,j+1).fill = FILL_HDR
            auto_col(ws,j+1,14)
        r += 1
        for _, row in df.iterrows():
            for j,v in enumerate(row):
                cell = ws.cell(r,j+1,value=v if not (isinstance(v,float) and np.isnan(v)) else 0)
                cell.font = FN; cell.alignment = AR; cell.number_format = NF
            r += 1
        # Heatmap
        _apply_heatmap(ws, r-len(df), r-1, 2, len(df.columns)-1)
        r += 2
    fit_to_width(ws)


def _apply_heatmap(ws, start_row, end_row, start_col, num_cols):
    vals = []
    for r in range(start_row, end_row+1):
        for c in range(start_col, start_col+num_cols):
            v = ws.cell(r,c).value
            if isinstance(v,(int,float)): vals.append(v)
    if not vals: return
    mn,mx = min(vals),max(vals); rng = mx-mn if mx!=mn else 1
    for r in range(start_row, end_row+1):
        for c in range(start_col, start_col+num_cols):
            v = ws.cell(r,c).value
            if isinstance(v,(int,float)):
                pct = (v-mn)/rng
                red = int(255*(1-pct)); green = int(200*pct+55)
                ws.cell(r,c).fill = PatternFill("solid",fgColor=f"{min(red,255):02X}{min(green,255):02X}55")


def build_monte_carlo(wb, mc_result, mc_config=None, n=5):
    """Write formula-linked Monte Carlo simulation tab.

    Creates 1,000 live iterations using NORM.INV(RAND(),...) formulas
    linked to the Assumptions and WACC sheets.  Press F9 in Excel to
    resimulate with fresh random draws.
    """
    ws = wb.create_sheet("Monte Carlo")
    ws.sheet_properties.tabColor = "7030A0"

    if mc_config is None and mc_result is None:
        ws.cell(1, 1, value="No Monte Carlo data").font = FN
        return

    # Fallback: static output when no config is available (should not
    # happen in normal operation since cfg.monte_carlo always exists)
    if mc_config is None:
        ws.cell(1, 1, value="No Monte Carlo config — static data only").font = FN
        fit_to_width(ws)
        return

    ITERS = 1000  # capped for formula-based simulation

    # ── Column widths ────────────────────────────────────────────
    auto_col(ws, 1, 24)
    auto_col(ws, 2, 16)
    auto_col(ws, 3, 16)

    # ── Title ────────────────────────────────────────────────────
    ws.cell(1, 1, value="Monte Carlo Simulation (Live Excel)").font = FK
    ws.cell(2, 1,
            value="Press F9 to resimulate  \u00b7  "
                  "All parameters auto-linked to Assumptions & WACC"
            ).font = Font(name="Verdana", size=8, italic=True, color="666666")

    # ── Section 1: Simulation Parameters (rows 3-23) ────────────
    ws.cell(3, 1, value="Simulation Parameters").font = FB

    # (row, label, value, is_formula, number_format, is_editable_input)
    _params = [
        (4,  "Revenue Growth \u03bc",
             f"={ar('rev_cagr')}", True, PF, False),
        (5,  "Revenue Growth \u03c3",
             mc_config.revenue_growth_std, False, PF, True),
        (6,  "EBITDA Margin \u03bc",
             f"=1-{ar('cogs_pct')}-{ar('sga_pct')}-{ar('oopex_pct')}",
             True, PF, False),
        (7,  "EBITDA Margin \u03c3",
             mc_config.ebitda_margin_std, False, PF, True),
        (8,  "WACC \u03bc",
             f"=WACC!$B${WR['WACC']}", True, PF, False),
        (9,  "WACC \u03c3",
             mc_config.wacc_std, False, PF, True),
        (10, "Terminal Growth \u03bc",
             f"={ar('tg')}", True, PF, False),
        (11, "Terminal Growth \u03c3",
             mc_config.terminal_growth_std, False, PF, True),
        (12, "Exit Multiple \u03bc",
             f"={ar('exit_m')}", True, XF, False),
        (13, "Exit Multiple \u03c3",
             mc_config.exit_multiple_std, False, XF, True),
        (14, "Base Revenue",
             f"={ar('brev')}", True, NF, False),
        (15, "Tax Rate",
             f"={ar('tax')}", True, PF, False),
        (16, "Capex % Revenue",
             f"={ar('capex_pct')}", True, PF, False),
        (17, "D&A Rate (% of Rev)",
             f"={ar('dep_rate')}", True, PF, False),
        (18, "Cash (Equity Bridge)",
             f"={ar('cash')}", True, NF, False),
        (19, "Debt (Equity Bridge)",
             f"={ar('debt')}", True, NF, False),
        (20, "Shares Outstanding",
             f"={ar('shares')}", True, NF, False),
        (21, "Gordon Weight",
             f"={ar('gw')}", True, PF, False),
        (22, "Projection Years",
             n, False, '#,##0', False),
        (23, "Iterations (Formula)",
             ITERS, False, '#,##0', False),
    ]

    for row, label, val, _is_formula, fmt, is_input in _params:
        ws.cell(row, 1, value=label).font = FN
        c = ws.cell(row, 2)
        c.value = val
        c.number_format = fmt
        if is_input:
            c.font = FI
            c.fill = FILL_INP
        else:
            c.font = FN

    # ── Pre-build formula fragments ──────────────────────────────
    # Inline array constants for SUMPRODUCT, e.g. {1,2,3,4,5}
    years_arr = ",".join(str(i) for i in range(1, n + 1))
    # Mid-year discount exponents, e.g. {0.5,1.5,2.5,3.5,4.5}
    midyear_arr = ",".join(str(round(i - 0.5, 1)) for i in range(1, n + 1))
    tv_disc_exp = round(n - 0.5, 1)  # e.g. 4.5 for n=5

    # Simulation table row range
    SIM_HDR = 66
    SIM_START = 67
    SIM_END = SIM_START + ITERS - 1   # 1066
    eq_range = f"$I${SIM_START}:$I${SIM_END}"  # equity values column

    # ── Section 2: Output Statistics (rows 25-39) ────────────────
    ws.cell(25, 1, value="Output Statistics (Live)").font = FB

    _stats = [
        (26, "Mean",             f"=AVERAGE({eq_range})", NF),
        (27, "Median",           f"=MEDIAN({eq_range})", NF),
        (28, "Std Dev",          f"=STDEV({eq_range})", NF),
        (29, "P10",              f"=PERCENTILE({eq_range},0.1)", NF),
        (30, "P25",              f"=PERCENTILE({eq_range},0.25)", NF),
        (31, "P50",              f"=PERCENTILE({eq_range},0.5)", NF),
        (32, "P75",              f"=PERCENTILE({eq_range},0.75)", NF),
        (33, "P90",              f"=PERCENTILE({eq_range},0.9)", NF),
        (34, "Min",              f"=MIN({eq_range})", NF),
        (35, "Max",              f"=MAX({eq_range})", NF),
        (36, "Per Share Mean",
             f"=IF($B$20=0,0,AVERAGE({eq_range})/$B$20)", PF2),
        (37, "Per Share Median",
             f"=IF($B$20=0,0,MEDIAN({eq_range})/$B$20)", PF2),
        (38, "Per Share P10",
             f"=IF($B$20=0,0,PERCENTILE({eq_range},0.1)/$B$20)", PF2),
        (39, "Per Share P90",
             f"=IF($B$20=0,0,PERCENTILE({eq_range},0.9)/$B$20)", PF2),
    ]
    for row, label, formula, fmt in _stats:
        ws.cell(row, 1, value=label).font = FN
        c = ws.cell(row, 2)
        c.value = formula
        c.font = FN
        c.number_format = fmt

    # ── Section 3: Histogram (rows 41-63) ────────────────────────
    HIST_BINS = 20
    ws.cell(41, 1, value="Distribution").font = FB

    # Helper cells for bin bounds (placed in cols E-J of row 41)
    ws.cell(41, 5, value="min:").font = Font(
        name="Verdana", size=8, color="999999")
    ws.cell(41, 6).value = f"=MIN({eq_range})"
    ws.cell(41, 6).number_format = NF
    ws.cell(41, 6).font = FN
    ws.cell(41, 7, value="max:").font = Font(
        name="Verdana", size=8, color="999999")
    ws.cell(41, 8).value = f"=MAX({eq_range})"
    ws.cell(41, 8).number_format = NF
    ws.cell(41, 8).font = FN
    ws.cell(41, 9, value="width:").font = Font(
        name="Verdana", size=8, color="999999")
    ws.cell(41, 10).value = f"=(H41-F41)/{HIST_BINS}"
    ws.cell(41, 10).number_format = NF
    ws.cell(41, 10).font = FN

    # Histogram headers
    ws.cell(42, 1, value="Bin Center").font = FH
    ws.cell(42, 1).fill = FILL_HDR
    ws.cell(42, 2, value="Frequency").font = FH
    ws.cell(42, 2).fill = FILL_HDR

    for bi in range(HIST_BINS):
        br = 43 + bi
        # Bin center = min + (bi + 0.5) * width
        ws.cell(br, 1).value = f"=$F$41+({bi}+0.5)*$J$41"
        ws.cell(br, 1).number_format = NF
        ws.cell(br, 1).font = FN
        # Frequency via COUNTIFS
        lo = f"$F$41+{bi}*$J$41"
        hi = f"$F$41+{bi + 1}*$J$41"
        if bi < HIST_BINS - 1:
            ws.cell(br, 2).value = (
                f'=COUNTIFS({eq_range},">="&({lo}),'
                f'{eq_range},"<"&({hi}))'
            )
        else:
            ws.cell(br, 2).value = f'=COUNTIFS({eq_range},">="&({lo}))'
        ws.cell(br, 2).number_format = NF
        ws.cell(br, 2).font = FN

    hist_end = 43 + HIST_BINS - 1   # row 62

    # Bar chart of histogram
    chart = BarChart()
    chart.type = "col"
    chart.title = "Equity Value Distribution (Live \u2014 F9 to Refresh)"
    chart.y_axis.title = "Frequency"
    chart.x_axis.title = "Equity Value"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    data = Reference(ws, min_col=2, min_row=42, max_row=hist_end)
    cats = Reference(ws, min_col=1, min_row=43, max_row=hist_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "D3")

    # ── Section 4: Simulation Table (rows 65-1066) ───────────────
    ws.cell(65, 1,
            value=f"Simulation Table ({ITERS:,} Iterations "
                  f"\u2014 Press F9 to Resimulate)").font = FB

    # Column headers
    _hdrs = [
        ("#", 14), ("Rev Growth", 14), ("EBITDA Margin", 14),
        ("Terminal Growth", 14), ("WACC", 14), ("Exit Multiple", 14),
        ("PV of FCFs", 16), ("PV of TV", 16),
        ("Equity Value", 16), ("Per Share", 14),
    ]
    for j, (h, w) in enumerate(_hdrs, 1):
        ws.cell(SIM_HDR, j, value=h).font = FH
        ws.cell(SIM_HDR, j).fill = FILL_HDR
        auto_col(ws, j, w)

    # ── Write 1,000 formula-based iteration rows ─────────────────
    for i in range(ITERS):
        row = SIM_START + i

        # Col A: iteration number
        ws.cell(row, 1, value=i + 1).font = FN
        ws.cell(row, 1).number_format = '#,##0'

        # Col B: Revenue Growth ~ N(μ, σ) clipped [-20%, 50%]
        ws.cell(row, 2).value = (
            f'=MAX(-0.2,MIN(0.5,_xlfn.NORM.INV(RAND(),$B$4,$B$5)))'
        )
        ws.cell(row, 2).font = FN
        ws.cell(row, 2).number_format = PF

        # Col C: EBITDA Margin ~ N(μ, σ) clipped [1%, 60%]
        ws.cell(row, 3).value = (
            f'=MAX(0.01,MIN(0.6,_xlfn.NORM.INV(RAND(),$B$6,$B$7)))'
        )
        ws.cell(row, 3).font = FN
        ws.cell(row, 3).number_format = PF

        # Col D: Terminal Growth ~ N(μ, σ) clipped [0%, 5%]
        ws.cell(row, 4).value = (
            f'=MAX(0,MIN(0.05,_xlfn.NORM.INV(RAND(),$B$10,$B$11)))'
        )
        ws.cell(row, 4).font = FN
        ws.cell(row, 4).number_format = PF

        # Col E: WACC ~ N(μ, σ) clipped [3%, 30%], forced > TG+1%
        ws.cell(row, 5).value = (
            f'=MAX(D{row}+0.01,'
            f'MAX(0.03,MIN(0.3,_xlfn.NORM.INV(RAND(),$B$8,$B$9))))'
        )
        ws.cell(row, 5).font = FN
        ws.cell(row, 5).number_format = PF

        # Col F: Exit Multiple ~ N(μ, σ) clipped [3x, 25x]
        ws.cell(row, 6).value = (
            f'=MAX(3,MIN(25,_xlfn.NORM.INV(RAND(),$B$12,$B$13)))'
        )
        ws.cell(row, 6).font = FN
        ws.cell(row, 6).number_format = XF

        # Col G: PV of projected FCFs (SUMPRODUCT over n years)
        #   k = (margin − da_pct)·(1−tax) + da_pct − capex_pct
        #   FCF_yr = base_rev · (1+g)^yr · k
        #   PV_yr  = FCF_yr / (1+w)^(yr−0.5)
        pv_formula = (
            f"=SUMPRODUCT($B$14"
            f"*((C{row}-$B$17)*(1-$B$15)+$B$17-$B$16)"
            f"*(1+B{row})^{{{years_arr}}}"
            f"/((1+E{row})^{{{midyear_arr}}}))"
        )
        ws.cell(row, 7).value = pv_formula
        ws.cell(row, 7).font = FN
        ws.cell(row, 7).number_format = NF

        # Col H: PV of blended Terminal Value
        #   Gordon TV = last_FCF·(1+tg) / max(w−tg, 1e-6)
        #   Exit   TV = last_EBITDA · exit_mult
        #   PV TV     = (Gordon·gw + Exit·(1−gw)) / (1+w)^(n−0.5)
        tv_formula = (
            f"=(($B$14*(1+B{row})^{n}"
            f"*((C{row}-$B$17)*(1-$B$15)+$B$17-$B$16)"
            f"*(1+D{row})/MAX(E{row}-D{row},0.000001))*$B$21"
            f"+($B$14*(1+B{row})^{n}*C{row}*F{row})*(1-$B$21))"
            f"/((1+E{row})^{tv_disc_exp})"
        )
        ws.cell(row, 8).value = tv_formula
        ws.cell(row, 8).font = FN
        ws.cell(row, 8).number_format = NF

        # Col I: Equity Value = PV FCFs + PV TV + Cash − Debt
        ws.cell(row, 9).value = f"=G{row}+H{row}+$B$18-$B$19"
        ws.cell(row, 9).font = FN
        ws.cell(row, 9).number_format = NF

        # Col J: Per Share
        ws.cell(row, 10).value = f"=IF($B$20=0,0,I{row}/$B$20)"
        ws.cell(row, 10).font = FN
        ws.cell(row, 10).number_format = PF2

        # Alternate row shading for readability
        if i % 2 == 1:
            for col in range(1, 11):
                ws.cell(row, col).fill = FILL_ALT

    fit_to_width(ws)


def build_tornado(wb, tornado_result):
    ws = wb.create_sheet("Tornado")
    ws.sheet_properties.tabColor = "ED7D31"
    if tornado_result is None:
        ws.cell(1,1,value="No tornado data").font = FN; return
    ws.cell(1,1,value="Tornado Analysis").font = FK
    auto_col(ws,1,22)
    df = tornado_result.drivers
    r = 3
    for j,cn in enumerate(df.columns):
        ws.cell(r-1,j+1,value=cn).font = FH; ws.cell(r-1,j+1).fill = FILL_HDR
        auto_col(ws,j+1,16)
    for _,row in df.iterrows():
        for j,(cn,v) in enumerate(row.items()):
            cell = ws.cell(r,j+1,value=v)
            cell.font = FN; cell.number_format = NF if isinstance(v,(int,float)) else '@'
        r += 1
    chart = BarChart(); chart.type = "bar"
    chart.title = "Impact on Equity Value"; chart.style = 10
    chart.width = 20; chart.height = 12
    low = Reference(ws,min_col=5,min_row=2,max_row=r-1)
    high = Reference(ws,min_col=6,min_row=2,max_row=r-1)
    cats = Reference(ws,min_col=1,min_row=3,max_row=r-1)
    chart.add_data(low,titles_from_data=True); chart.add_data(high,titles_from_data=True)
    chart.set_categories(cats); ws.add_chart(chart,f"A{r+1}")
    fit_to_width(ws)


def build_comps(wb, comps_result):
    ws = wb.create_sheet("Comps")
    ws.sheet_properties.tabColor = "548235"
    if comps_result is None or comps_result.peer_table.empty:
        ws.cell(1,1,value="No comparable company data").font = FN; return
    ws.cell(1,1,value="Comparable Companies").font = FK
    pt = comps_result.peer_table
    auto_col(ws,1,14)
    for j,cn in enumerate(pt.columns,1):
        ws.cell(2,j,value=cn).font = FH; ws.cell(2,j).fill = FILL_HDR; auto_col(ws,j,16)
    for i,(_,row) in enumerate(pt.iterrows(),3):
        for j,(cn,v) in enumerate(row.items(),1):
            cell = ws.cell(i,j,value=v if v is not None and not (isinstance(v,float) and np.isnan(v)) else 0)
            cell.font = FN; cell.number_format = XF if cn != "Ticker" else '@'
    # Summary stats
    if hasattr(comps_result,'summary_stats') and not comps_result.summary_stats.empty:
        r = len(pt)+4
        ws.cell(r,1,value="Summary Statistics").font = FK; r += 1
        ss = comps_result.summary_stats
        for j,cn in enumerate(ss.columns,1):
            ws.cell(r,j,value=cn).font = FH; ws.cell(r,j).fill = FILL_HDR
        r += 1
        for _,row in ss.iterrows():
            for j,(cn,v) in enumerate(row.items(),1):
                ws.cell(r,j,value=v if v is not None else 0).font = FN
                ws.cell(r,j).number_format = XF if cn != "Statistic" else '@'
            r += 1
    fit_to_width(ws)


def build_checks(wb, n):
    """Formula-linked checks tab."""
    ws = wb.create_sheet("Checks")
    ws.sheet_properties.tabColor = "00B050"
    auto_col(ws,1,28)
    ws.cell(1,1,value="Model Integrity Checks").font = FK
    for j in range(n):
        ws.cell(2,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,14)
    hdr_row(ws,2,n+1)
    style_label(ws,3,1,"BS Balance (A-L-E)",bold=True)
    style_label(ws,4,1,"BS Status")
    style_label(ws,6,1,"CF Ending Cash = BS Cash",bold=True)
    style_label(ws,7,1,"CF Status")
    for j in range(n):
        c = CL(j+2)
        ws[f"{c}3"] = f"='Balance Sheet'!{c}{BS['Chk']}"
        ws[f"{c}4"] = f"='Balance Sheet'!{c}{BS['Stat']}"
        ws[f"{c}6"] = f"='Cash Flow'!{c}23-'Balance Sheet'!{c}{BS['Cash']}"
        ws[f"{c}7"] = f'=IF(ABS({c}6)<1,"PASS","FAIL")'
        for r in [3,6]:
            ws.cell(r,j+2).number_format = NF; ws.cell(r,j+2).font = FN
        for r in [4,7]:
            ws.cell(r,j+2).font = FN
    fit_to_width(ws)


def build_audit(wb, cfg):
    ws = wb.create_sheet("Audit Trail")
    ws.sheet_properties.tabColor = "A5A5A5"
    auto_col(ws,1,24); auto_col(ws,2,18); auto_col(ws,3,30)
    ws.cell(1,1,value="Audit Trail").font = FK
    ws.cell(3,1,value="Generated At").font = FB
    ws.cell(3,2,value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font = FN
    ws.cell(4,1,value="Company").font = FB
    ws.cell(4,2,value=cfg.company.name).font = FN
    r = 6
    ws.cell(r,1,value="Configuration Snapshot").font = FK; r += 1
    try:
        d = cfg.to_dict()
        for section,params in d.items():
            if isinstance(params,dict):
                for k,v in params.items():
                    ws.cell(r,1,value=section).font = FN
                    ws.cell(r,2,value=k).font = FN
                    ws.cell(r,3,value=str(v)).font = FN; r += 1
    except Exception:
        ws.cell(r,1,value="Config serialization not available").font = FN
    fit_to_width(ws)
