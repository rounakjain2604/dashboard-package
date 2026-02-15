"""Core formula-linked sheets: Cover, Assumptions, IS, WC, CapexDA, Debt."""
from __future__ import annotations
from datetime import datetime
from openpyxl.utils import get_column_letter as CL
from .excel_formats import *

# ── Row constants for cross-sheet references ──────────────────────────
# Income Statement rows
IS = {"Rev":2,"Gro":3,"COGS":4,"GP":5,"GM":6,"SGA":7,"OOp":8,
      "EBITDA":9,"EM":10,"Dep":11,"Amort":12,"DA":13,"EBIT":14,
      "Int":15,"EBT":16,"Tax":17,"NI":18}
# Working Capital rows
WC = {"Rev":2,"COGS":3,"AR":4,"Inv":6,"Prep":8,"OCA":9,
      "AP":10,"Accr":12,"OCL":13,"NWC":14,"dNWC":15}
# Capex DA rows
CD = {"Beg":2,"Capex":3,"DepEx":4,"DepNew":5,"Dep":6,"End":7}
# Debt Schedule rows
DS = {"Beg":2,"New":3,"Int":4,"Rep":5,"End":6,"Cur":7}
# Balance Sheet rows
BS = {"Cash":2,"AR":3,"Inv":4,"Prep":5,"OCA":6,"CA":7,
      "PPE":8,"GW":9,"Intan":10,"OLA":11,"NCA":12,"TA":13,
      "AP":15,"Accr":16,"OCL":17,"CurD":18,"CL":19,
      "LTD":20,"OLL":21,"NCL":22,"TL":23,
      "CS":25,"RE":26,"TE":27,"TLE":28,"Chk":30,"Stat":31}
# Cash Flow rows
CF = {"NI":3,"Dep":4,"Amort":5,"DA":6,"dNWC":7,"CFO":8,
      "Capex":11,"CFI":12,
      "NewD":15,"RepD":16,"IntP":17,"Div":18,"CFF":19,
      "Net":21,"BegC":22,"EndC":23,
      "EBIT":26,"NOPAT":27,"DA2":28,"Capex2":29,"dNWC2":30,"UFCF":31}
# WACC rows
WR = {"Rf":4,"Beta":5,"ERP":6,"SP":7,"CRP":8,"Ke":9,
      "KdPre":12,"KdPost":13,"Ew":16,"Dw":17,"WACC":19}
# DCF rows
DC = {"UFCF":2,"DF":3,"PV":4,"SumPV":6,
      "TFCF":9,"TEBITDA":10,"TVG":11,"TVE":12,"TVB":13,"PVTV":14,
      "EV":17,"Cash":18,"Debt":19,"Equity":20,"Shares":21,"Price":22}


def build_cover(wb, cfg):
    ws = wb.create_sheet("Cover")
    ws.sheet_properties.tabColor = "1F3864"
    ws.merge_cells("A6:F6")
    ws.cell(6,1,value=cfg.company.name).font = Font(name="Verdana",size=24,bold=True,color="1F3864")
    ws.merge_cells("A8:F8")
    ws.cell(8,1,value="Discounted Cash Flow Valuation").font = FT
    for r,lv in [(11,("Ticker",cfg.company.ticker)),(12,("Industry",cfg.company.industry)),
                  (13,("Analyst",cfg.company.analyst_name)),(14,("Date",cfg.company.report_date or datetime.now().strftime("%Y-%m-%d")))]:
        ws.cell(r,1,value=lv[0]).font = FB
        ws.cell(r,2,value=lv[1]).font = FN
    ws.cell(16,1,value="CONFIDENTIAL").font = Font(name="Verdana",size=8,color="999999")
    auto_col(ws,1,20); auto_col(ws,2,30)
    fit_to_width(ws)


def build_assumptions(wb, cfg, base_rev, base_cash, base_ppe, base_nwc, base_re, base_cs, credit_spread=0.02):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = "4472C4"
    auto_col(ws,1,28); auto_col(ws,2,18)
    f = cfg.forecast; w = cfg.wacc; v = cfg.valuation

    def _inp(row, label, val, fmt=NF):
        ws.cell(row,1,value=label).font = FN
        c = ws.cell(row,2,value=val)
        c.font = FI; c.fill = FILL_INP; c.number_format = fmt

    def _sec(row, label):
        ws.cell(row,1,value=label).font = FK
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)

    _sec(2, "FORECAST ASSUMPTIONS")
    _inp(A["rev_cagr"], "Revenue CAGR", f.revenue_cagr, PF)
    _inp(A["cogs_pct"], "COGS % Revenue", f.cogs_pct_revenue, PF)
    _inp(A["sga_pct"], "SGA % Revenue", f.sga_pct_revenue, PF)
    _inp(A["oopex_pct"], "Other OpEx % Revenue", f.other_opex_pct_revenue, PF)
    _inp(A["dep_rate"], "Depreciation Rate", f.depreciation_rate, PF)
    _inp(A["capex_pct"], "Capex % Revenue", f.capex_pct_revenue, PF)
    _inp(A["tax"], "Tax Rate", f.tax_rate, PF)
    _inp(A["dso"], "DSO (days)", f.dso, NF)
    _inp(A["dio"], "DIO (days)", f.dio, NF)
    _inp(A["dpo"], "DPO (days)", f.dpo, NF)
    _inp(A["div"], "Dividend Payout Ratio", f.dividend_payout_ratio, PF)
    _inp(A["amort_pct"], "Amortisation % Revenue", f.amortisation_pct_revenue, PF)
    _inp(A["prepaid_pct"], "Prepaid % Revenue", f.prepaid_pct_revenue, PF)
    _inp(A["accrued_pct"], "Accrued % Revenue", f.accrued_pct_revenue, PF)
    _inp(A["oca_pct"], "Other CA % Revenue", f.other_current_assets_pct_revenue, PF)
    _inp(A["ocl_pct"], "Other CL % Revenue", f.other_current_liabilities_pct_revenue, PF)

    _sec(21, "WACC ASSUMPTIONS")
    _inp(A["rf"], "Risk-Free Rate", w.risk_free_rate, PF)
    _inp(A["erp"], "Equity Risk Premium", w.equity_risk_premium, PF)
    _inp(A["beta"], "Beta", w.beta, '0.00')
    _inp(A["sp"], "Size Premium", w.size_premium, PF)
    _inp(A["crp"], "Country Risk Premium", w.country_risk_premium, PF)
    _inp(A["dw"], "Debt Weight", w.target_debt_weight, PF)
    _inp(A["ew"], "Equity Weight", w.target_equity_weight, PF)
    _inp(A["icr"], "Interest Coverage Ratio", w.interest_coverage_ratio, '0.0x')
    ws.cell(31,1,value="Credit Spread (computed)").font = FN
    ws.cell(31,2,value=credit_spread).font = FN; ws.cell(31,2).number_format = PF

    _sec(33, "VALUATION ASSUMPTIONS")
    _inp(A["tg"], "Terminal Growth Rate", v.terminal_growth_rate, PF)
    _inp(A["exit_m"], "Exit EV/EBITDA Multiple", v.exit_ev_ebitda_multiple, '0.0x')
    _inp(A["gw"], "Gordon Growth Weight", v.gordon_weight, PF)
    _inp(A["shares"], "Shares Outstanding (M)", v.fully_diluted_shares/1e6 if v.fully_diluted_shares>1e6 else v.fully_diluted_shares, NF)
    _inp(A["cash"], "Cash (Equity Bridge)", v.cash, NF)
    _inp(A["debt"], "Debt (Equity Bridge)", v.debt, NF)

    _sec(41, "BASE YEAR INPUTS")
    _inp(A["brev"], "Base Year Revenue", base_rev, NF)
    _inp(A["bcash"], "Base Year Cash", base_cash, NF)
    _inp(A["bppe"], "Base Year PP&E (Net)", base_ppe, NF)
    _inp(A["bnwc"], "Base Year NWC", base_nwc, NF)
    _inp(A["bre"], "Base Retained Earnings", base_re, NF)
    _inp(A["bcs"], "Base Common Stock", base_cs, NF)

    # ── Debt Schedule Inputs ─────────────────────────────────────────
    _sec(49, "DEBT SCHEDULE INPUTS")
    # Use first tranche data if available, otherwise defaults
    _debt_bal = 0.0
    _debt_rate = 0.06
    _debt_amort = 0.0
    _debt_maturity = 5
    if cfg.debt_tranches:
        t = cfg.debt_tranches[0]
        _debt_bal = t.beginning_balance
        _debt_rate = t.interest_rate
        _debt_amort = t.annual_amortisation
        _debt_maturity = t.maturity_year
    elif v.debt > 0:
        _debt_bal = v.debt
    _inp(A["debt_bal"], "Debt Beginning Balance", _debt_bal, NF)
    _inp(A["debt_rate"], "Debt Interest Rate", _debt_rate, PF)
    _inp(A["debt_amort"], "Debt Annual Amortisation", _debt_amort, NF)
    _inp(A["debt_maturity"], "Debt Maturity (Year)", _debt_maturity, '#,##0')

    # ── Other Balance Sheet Items ────────────────────────────────────
    _sec(55, "OTHER BALANCE SHEET ITEMS")
    _inp(A["goodwill"], "Goodwill", 0, NF)
    _inp(A["intangibles"], "Intangibles (Net)", 0, NF)
    _inp(A["olt_assets"], "Other Long-Term Assets", 0, NF)
    _inp(A["olt_liab"], "Other Long-Term Liabilities", 0, NF)
    fit_to_width(ws)


def build_is(wb, n):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_properties.tabColor = "2E75B6"
    auto_col(ws,1,22)
    # Headers
    ws.cell(1,1,value="Income Statement").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}")
        auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    # Labels
    items = ["Revenue","Revenue Growth","COGS","Gross Profit","Gross Margin",
             "SGA","Other OpEx","EBITDA","EBITDA Margin",
             "Depreciation","Amortisation","Total D&A","EBIT",
             "Interest Expense","EBT","Tax Expense","Net Income"]
    pct_rows = {IS["Gro"],IS["GM"],IS["EM"]}
    tot_rows = {IS["GP"],IS["EBITDA"],IS["EBIT"],IS["EBT"],IS["NI"]}
    for i,lbl in enumerate(items,2):
        style_label(ws,i,1,lbl,bold=(i in tot_rows))
    # Formulas
    for j in range(n):
        col = j+2; c = CL(col); pc = CL(col-1) if j>0 else None
        if j==0:
            ws[f"{c}{IS['Rev']}"] = f"={ar('brev')}*(1+{ar('rev_cagr')})"
        else:
            ws[f"{c}{IS['Rev']}"] = f"={pc}{IS['Rev']}*(1+{ar('rev_cagr')})"
        ws[f"{c}{IS['Gro']}"]  = f"={ar('rev_cagr')}"
        ws[f"{c}{IS['COGS']}"] = f"={c}{IS['Rev']}*{ar('cogs_pct')}"
        ws[f"{c}{IS['GP']}"]   = f"={c}{IS['Rev']}-{c}{IS['COGS']}"
        ws[f"{c}{IS['GM']}"]   = f"=IF({c}{IS['Rev']}=0,0,{c}{IS['GP']}/{c}{IS['Rev']})"
        ws[f"{c}{IS['SGA']}"]  = f"={c}{IS['Rev']}*{ar('sga_pct')}"
        ws[f"{c}{IS['OOp']}"]  = f"={c}{IS['Rev']}*{ar('oopex_pct')}"
        ws[f"{c}{IS['EBITDA']}"]= f"={c}{IS['GP']}-{c}{IS['SGA']}-{c}{IS['OOp']}"
        ws[f"{c}{IS['EM']}"]   = f"=IF({c}{IS['Rev']}=0,0,{c}{IS['EBITDA']}/{c}{IS['Rev']})"
        ws[f"{c}{IS['Dep']}"]  = f"='{sr('Capex DA',c,CD['Dep'])[1:]}"[:-1] if False else f"='Capex DA'!{c}{CD['Dep']}"
        ws[f"{c}{IS['Amort']}"]= f"={c}{IS['Rev']}*{ar('amort_pct')}"
        ws[f"{c}{IS['DA']}"]   = f"={c}{IS['Dep']}+{c}{IS['Amort']}"
        ws[f"{c}{IS['EBIT']}"] = f"={c}{IS['EBITDA']}-{c}{IS['DA']}"
        ws[f"{c}{IS['Int']}"]  = f"='Debt Schedule'!{c}{DS['Int']}"
        ws[f"{c}{IS['EBT']}"]  = f"={c}{IS['EBIT']}-{c}{IS['Int']}"
        ws[f"{c}{IS['Tax']}"]  = f"=MAX({c}{IS['EBT']}*{ar('tax')},0)"
        ws[f"{c}{IS['NI']}"]   = f"={c}{IS['EBT']}-{c}{IS['Tax']}"
        # Formatting
        for r in range(2,19):
            fmt_cells(ws,r,col,col, PF if r in pct_rows else NF, bold=(r in tot_rows))
    fit_to_width(ws)


def build_wc(wb, n):
    ws = wb.create_sheet("Working Capital")
    ws.sheet_properties.tabColor = "548235"
    auto_col(ws,1,24)
    ws.cell(1,1,value="Working Capital").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    items = {2:"Revenue (ref)",3:"COGS (ref)",4:"Accounts Receivable",5:"DSO",
             6:"Inventory",7:"DIO",8:"Prepaid Expenses",9:"Other Current Assets",
             10:"Accounts Payable",11:"DPO",12:"Accrued Liabilities",
             13:"Other Current Liabilities",14:"Net Working Capital",15:"Delta NWC"}
    for r,lbl in items.items():
        style_label(ws,r,1,lbl,bold=(r in {14,15}))
    for j in range(n):
        col=j+2; c=CL(col); pc=CL(col-1) if j>0 else None
        ws[f"{c}{WC['Rev']}"]  = f"='Income Statement'!{c}{IS['Rev']}"
        ws[f"{c}{WC['COGS']}"] = f"='Income Statement'!{c}{IS['COGS']}"
        ws[f"{c}4"]  = f"={c}{WC['Rev']}*{ar('dso')}/365"
        ws[f"{c}5"]  = f"={ar('dso')}"
        ws[f"{c}6"]  = f"={c}{WC['COGS']}*{ar('dio')}/365"
        ws[f"{c}7"]  = f"={ar('dio')}"
        ws[f"{c}8"]  = f"={c}{WC['Rev']}*{ar('prepaid_pct')}"
        ws[f"{c}9"]  = f"={c}{WC['Rev']}*{ar('oca_pct')}"
        ws[f"{c}10"] = f"={c}{WC['COGS']}*{ar('dpo')}/365"
        ws[f"{c}11"] = f"={ar('dpo')}"
        ws[f"{c}12"] = f"={c}{WC['Rev']}*{ar('accrued_pct')}"
        ws[f"{c}13"] = f"={c}{WC['Rev']}*{ar('ocl_pct')}"
        ws[f"{c}14"] = f"={c}4+{c}6+{c}8+{c}9-{c}10-{c}12-{c}13"
        if j==0:
            ws[f"{c}15"] = f"={c}14-{ar('bnwc')}"
        else:
            ws[f"{c}15"] = f"={c}14-{pc}14"
        for r in range(2,16):
            fmt_cells(ws,r,col,col, NF if r not in {5,7,11} else '#,##0', bold=(r in {14,15}))
    fit_to_width(ws)


def build_capex_da(wb, n):
    ws = wb.create_sheet("Capex DA")
    ws.sheet_properties.tabColor = "ED7D31"
    auto_col(ws,1,24)
    ws.cell(1,1,value="Capex & Depreciation").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    for r,lbl in {2:"Beg PP&E (Net)",3:"Capex",4:"Dep on Existing",5:"Dep on New Capex",6:"Total Depreciation",7:"Ending PP&E (Net)"}.items():
        style_label(ws,r,1,lbl,bold=(r in {6,7}))
    for j in range(n):
        col=j+2; c=CL(col); pc=CL(col-1) if j>0 else None
        if j==0:
            ws[f"{c}2"] = f"={ar('bppe')}"
        else:
            ws[f"{c}2"] = f"={pc}7"
        ws[f"{c}3"] = f"='Income Statement'!{c}{IS['Rev']}*{ar('capex_pct')}"
        ws[f"{c}4"] = f"={c}2*{ar('dep_rate')}"
        ws[f"{c}5"] = f"={c}3*{ar('dep_rate')}*0.5"
        ws[f"{c}6"] = f"={c}4+{c}5"
        ws[f"{c}7"] = f"={c}2+{c}3-{c}6"
        for r in range(2,8):
            fmt_cells(ws,r,col,col, NF, bold=(r in {6,7}))
    fit_to_width(ws)


def build_debt_schedule(wb, n, debt_values=None):
    """Write debt schedule using Excel formulas referencing Assumptions inputs."""
    ws = wb.create_sheet("Debt Schedule")
    ws.sheet_properties.tabColor = "A5A5A5"
    auto_col(ws,1,24)
    ws.cell(1,1,value="Debt Schedule").font = FK
    for j in range(n):
        ws.cell(1,j+2,value=f"Year {j+1}"); auto_col(ws,j+2,15)
    hdr_row(ws,1,n+1)
    for r,lbl in {2:"Beginning Balance",3:"New Issuance",4:"Interest Expense",5:"Principal Repayment",6:"Ending Balance",7:"Current Portion"}.items():
        style_label(ws,r,1,lbl,bold=(r in {4,6}))
    # All formulas reference Assumptions sheet inputs
    for j in range(n):
        col = j+2; c = CL(col); pc = CL(col-1) if j > 0 else None
        yr = j + 1  # 1-based year index
        # Beginning Balance: Year 1 = Assumptions debt_bal, Year 2+ = prev Ending
        if j == 0:
            ws[f"{c}{DS['Beg']}"] = f"={ar('debt_bal')}"
        else:
            ws[f"{c}{DS['Beg']}"] = f"={pc}{DS['End']}"
        # New Issuance = 0 (no new debt assumed)
        ws[f"{c}{DS['New']}"] = 0
        # Interest Expense = Beginning Balance * Interest Rate
        ws[f"{c}{DS['Int']}"] = f"={c}{DS['Beg']}*{ar('debt_rate')}"
        # Principal Repayment = MIN(Amortisation, Beginning) + bullet at maturity
        # At maturity year, repay everything remaining; otherwise just amortisation
        ws[f"{c}{DS['Rep']}"] = (
            f"=IF({yr}={ar('debt_maturity')},"
            f"{c}{DS['Beg']},"
            f"MIN({ar('debt_amort')},{c}{DS['Beg']}))"
        )
        # Ending Balance = Beginning - Repayment
        ws[f"{c}{DS['End']}"] = f"=MAX({c}{DS['Beg']}-{c}{DS['Rep']},0)"
        # Current Portion = MIN(Amortisation, Ending) — what's due next year
        ws[f"{c}{DS['Cur']}"] = (
            f"=IF({yr}>={ar('debt_maturity')},"
            f"{c}{DS['End']},"
            f"MIN({ar('debt_amort')},{c}{DS['End']}))"
        )
        for r in range(2,8):
            fmt_cells(ws,r,col,col, NF, bold=(r in {4,6}))
    fit_to_width(ws)
